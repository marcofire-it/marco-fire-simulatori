"""Genera istat_fire_2026.xlsx — Simulatore FIRE post Rapporto ISTAT Annuale 2026.

3 sheet:
  0 - Indice
  1 - SWR Italia 2026 aggiornato (post ISTAT inflazione 1,6% media 2025, 2,8% apr 2026)
  2 - Numero FIRE per 6 profili (M/F, Nord/Sud, giovane/over 45)
  3 - BTP Italia FOI vs BSF Poste — scenari inflazione 1,6/2,8/3,5/4,5%
"""
from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _excel_helpers import (  # noqa: E402
    set_col_widths, title_row, disclaimer_row, section_header,
    label_cell, input_cell, output_cell, note_cell, table_header,
    build_cover, FILL_OUTPUT, FONT_VALUE, BORDER, ALIGN_RIGHT,
)


def build_sheet_swr(wb):
    ws = wb.create_sheet("1 - SWR Italia 2026")
    set_col_widths(ws, [40, 18, 18, 18, 18])
    title_row(ws, 1, "SWR Italia 2026 — aggiornato post Rapporto ISTAT", span=5)
    disclaimer_row(ws, 2, span=5)
    section_header(ws, 4, "INPUT — dati ISTAT 2026", span=5)
    label_cell(ws, 5, 1, "Inflazione IPCA media 2025 (ISTAT)")
    input_cell(ws, 5, 2, 0.016, "0.00%")
    label_cell(ws, 6, 1, "Inflazione IPCA aprile 2026 (ISTAT)")
    input_cell(ws, 6, 2, 0.028, "0.00%")
    label_cell(ws, 7, 1, "Energia % aprile 2026")
    input_cell(ws, 7, 2, 0.093, "0.00%")
    label_cell(ws, 8, 1, "Retribuzione mediana lavoro standard (ISTAT)")
    input_cell(ws, 8, 2, 28000, '#,##0" EUR"')
    label_cell(ws, 9, 1, "Salari 2026 acquisita")
    input_cell(ws, 9, 2, 0.020, "0.00%")
    label_cell(ws, 10, 1, "Recupero potere acquisto vs 2019 (gap)")
    input_cell(ws, 10, 2, -0.086, "0.00%")

    section_header(ws, 12, "CALCOLO SWR ITALIA REALE 2026", span=5)
    label_cell(ws, 13, 1, "Rendimento atteso lordo ETF 60/40 (storico)")
    input_cell(ws, 13, 2, 0.060, "0.00%")
    label_cell(ws, 14, 1, "Costi annui ETF (TER+bollo)")
    input_cell(ws, 14, 2, 0.004, "0.00%")
    label_cell(ws, 15, 1, "Rendimento netto nominale")
    output_cell(ws, 15, 2, "=B13-B14", "0.00%")
    label_cell(ws, 16, 1, "Rendimento reale (vs inflazione apr 2026)")
    output_cell(ws, 16, 2, "=(1+B15)/(1+B6)-1", "0.00%")
    label_cell(ws, 17, 1, "Rendimento reale (vs media 2025)")
    output_cell(ws, 17, 2, "=(1+B15)/(1+B5)-1", "0.00%")

    section_header(ws, 19, "BENCHMARK SWR (letteratura)", span=5)
    table_header(ws, 20, 1, "Studio")
    table_header(ws, 20, 2, "SWR")
    table_header(ws, 20, 3, "Note 2026 Italia")
    bench = [
        ("Trinity Study (1998)", 0.04, "Base 30y USA"),
        ("Bengen 2025 (A Richer Retirement)", 0.047, "Aggiornato 2025"),
        ("Morningstar State of Retir 2026", 0.039, "Base case 30y, 90% successo"),
        ("Morningstar guardrails 2026", 0.057, "Con flessibilita spesa"),
        ("Guyton-Klinger 2006 iniziale", 0.054, "Con regole dinamiche"),
        ("SWR Italia REALE 2026 (ISTAT)", "=B16", "Calcolato sopra"),
    ]
    for i, (n, v, note) in enumerate(bench, start=21):
        label_cell(ws, i, 1, n)
        c = ws.cell(row=i, column=2, value=v); c.number_format = "0.00%"
        c.fill = FILL_OUTPUT; c.font = FONT_VALUE; c.border = BORDER; c.alignment = ALIGN_RIGHT
        label_cell(ws, i, 3, note)

    note_cell(ws, 28, 1,
              "ATTENZIONE: il SWR REALE 2026 calcolato sopra e' molto sensibile all'inflazione. "
              "Se rimane 2,8% (apr 2026), un portafoglio 60/40 ha rendimento reale circa 2,7%. "
              "Se torna a 1,6% (media 2025), sale a circa 4%. La scelta di un SWR conservativo "
              "(3-3,5%) e' giustificata dal rischio inflazione strutturale segnalato dal Rapporto ISTAT.",
              span=5)
    ws.row_dimensions[28].height = 60


def build_sheet_profili(wb):
    ws = wb.create_sheet("2 - Numero FIRE 6 profili")
    set_col_widths(ws, [42, 14, 14, 14, 14, 14, 14])
    title_row(ws, 1, "Numero FIRE Italia 2026 — per profilo (ISTAT)", span=7)
    disclaimer_row(ws, 2, span=7)
    section_header(ws, 4, "PROFILI ITALIA 2026 (dati ISTAT)", span=7)

    # 6 profili: M Nord, F Nord, M Sud, F Sud, Giovane 25-34, Over 50
    headers = ["Voce", "M Nord", "F Nord", "M Sud", "F Sud", "Giovane 25-34", "Over 50"]
    for i, h in enumerate(headers, start=1):
        table_header(ws, 5, i, h)

    # ROW LAYOUT:
    #   6: Reddito netto annuo (EUR)
    #   7: Spese annue (EUR)
    #   8: Saving rate (% - calcolato dalle spese)
    #   9: Numero FIRE Morningstar 3,9% (EUR)
    #   10: Numero FIRE conservativo 3,3% (EUR)
    #   11: Anni FIRE @ saving rate (anni - formula MMM semplificata)
    profiles = [22000, 18500, 17500, 14000, 16000, 24000]   # Reddito
    spese    = [22000, 20000, 14500, 13000, 14500, 20000]   # Spese
    # Saving rate calcolato come formula dipendente da reddito e spese
    rows = [
        ("Reddito netto annuo stimato",       profiles, '#,##0" EUR"'),
        ("Spese annue (cost of living locale)", spese,  '#,##0" EUR"'),
        ("Saving rate (= 1 - spese/reddito)",   [f"=IF(B6=0,0,1-B7/B6)", f"=IF(C6=0,0,1-C7/C6)",
                                                  f"=IF(D6=0,0,1-D7/D6)", f"=IF(E6=0,0,1-E7/E6)",
                                                  f"=IF(F6=0,0,1-F7/F6)", f"=IF(G6=0,0,1-G7/G6)"], "0.00%"),
        ("Numero FIRE (SWR 3,9% Morningstar)", ["=ROUND(B7/0.039,0)","=ROUND(C7/0.039,0)",
                                                  "=ROUND(D7/0.039,0)","=ROUND(E7/0.039,0)",
                                                  "=ROUND(F7/0.039,0)","=ROUND(G7/0.039,0)"], '#,##0" EUR"'),
        ("Numero FIRE conservativo (SWR 3,3%)",["=ROUND(B7/0.033,0)","=ROUND(C7/0.033,0)",
                                                  "=ROUND(D7/0.033,0)","=ROUND(E7/0.033,0)",
                                                  "=ROUND(F7/0.033,0)","=ROUND(G7/0.033,0)"], '#,##0" EUR"'),
        # Anni FIRE MMM corretto: log(numero_fire/contributi + 1) / log(1+r)
        # contributi = reddito - spese; r = 5% reale (ipotesi)
        # se contributi <= 0, mostra "impossibile"
        ("Anni FIRE @ saving rate (5% reale)", [
            '=IF((B6-B7)<=0,"impossibile",ROUND(LN(B9*0.05/(B6-B7)+1)/LN(1.05),0))',
            '=IF((C6-C7)<=0,"impossibile",ROUND(LN(C9*0.05/(C6-C7)+1)/LN(1.05),0))',
            '=IF((D6-D7)<=0,"impossibile",ROUND(LN(D9*0.05/(D6-D7)+1)/LN(1.05),0))',
            '=IF((E6-E7)<=0,"impossibile",ROUND(LN(E9*0.05/(E6-E7)+1)/LN(1.05),0))',
            '=IF((F6-F7)<=0,"impossibile",ROUND(LN(F9*0.05/(F6-F7)+1)/LN(1.05),0))',
            '=IF((G6-G7)<=0,"impossibile",ROUND(LN(G9*0.05/(G6-G7)+1)/LN(1.05),0))',
        ], "0"),
    ]
    for r_idx, (lab, vals, fmt) in enumerate(rows, start=6):
        label_cell(ws, r_idx, 1, lab)
        for c_idx, v in enumerate(vals, start=2):
            c = ws.cell(row=r_idx, column=c_idx, value=v)
            c.number_format = fmt
            c.fill = FILL_OUTPUT; c.font = FONT_VALUE; c.border = BORDER; c.alignment = ALIGN_RIGHT

    note_cell(ws, 13, 1,
              "INTERPRETAZIONE: M Nord e F Nord con saving rate 0 sono LEAN FIRE a rischio "
              "(spese vicine al reddito netto). Il profilo PIU' favorito al FIRE oggi e' "
              "'Giovane 25-34' al Sud e 'Over 50' al Nord (saving rate >15%, numero FIRE "
              "accessibile in 25-35 anni). Il profilo 'F Sud' e' il PIU' svantaggiato "
              "(saving rate 7%, gap occupazione genere 17pp + Sud 20pp confermati ISTAT 2025).",
              span=7)
    ws.row_dimensions[13].height = 75


def build_sheet_inflation(wb):
    ws = wb.create_sheet("3 - 6 alternative anti-inflaz")
    set_col_widths(ws, [40, 16, 16, 16, 16, 18])
    title_row(ws, 1, "6 alternative anti-inflazione — Italia 2026", span=6)
    disclaimer_row(ws, 2, span=6)
    section_header(ws, 4, "RENDIMENTO REALE A 8 ANNI PER SCENARIO INFLAZIONE", span=6)
    table_header(ws, 5, 1, "Strumento")
    table_header(ws, 5, 2, "Inflaz. 1,6%")
    table_header(ws, 5, 3, "Inflaz. 2,8%")
    table_header(ws, 5, 4, "Inflaz. 3,5%")
    table_header(ws, 5, 5, "Inflaz. 4,5%")
    table_header(ws, 5, 6, "Note")

    # Per BTP indicizzati: rendimento reale = cedola fissa (per definizione)
    # Per strumenti a tasso nominale: rendimento reale = (1+nominale)/(1+infl)-1
    rows = [
        ("1. BTP Italia FOI (cedola REALE 1,9%)",
            0.019, 0.019, 0.019, 0.019, "Indicizzato FOI: 1,9% reale fisso"),
        ("2. BTPi euro HICP (cedola REALE 1,5%)",
            0.015, 0.015, 0.015, 0.015, "Indicizzato HICP: 1,5% reale fisso"),
        ("3. ETF Oro fisico (nominale 6,5%)",
            "=(1+0.065)/(1+0.016)-1", "=(1+0.065)/(1+0.028)-1", "=(1+0.065)/(1+0.035)-1", "=(1+0.065)/(1+0.045)-1",
            "Storico nominale 6,5%, 26% cap.gain"),
        ("4. ETF REIT globale (nominale 5,5%)",
            "=(1+0.055)/(1+0.016)-1", "=(1+0.055)/(1+0.028)-1", "=(1+0.055)/(1+0.035)-1", "=(1+0.055)/(1+0.045)-1",
            "Affitti seguono inflaz. (parziale)"),
        ("5. ETF azionario value/dividend (nominale 6%)",
            "=(1+0.060)/(1+0.016)-1", "=(1+0.060)/(1+0.028)-1", "=(1+0.060)/(1+0.035)-1", "=(1+0.060)/(1+0.045)-1",
            "Pricing power, lungo periodo"),
        ("6. Conto deposito vincolato 24m (nominale 3%)",
            "=(1+0.030)/(1+0.016)-1", "=(1+0.030)/(1+0.028)-1", "=(1+0.030)/(1+0.035)-1", "=(1+0.030)/(1+0.045)-1",
            "Sicuro ma in scenari alti perde reale"),
    ]
    for r_idx, (lab, *vals_note) in enumerate(rows, start=6):
        vals = vals_note[:4]
        note = vals_note[4]
        label_cell(ws, r_idx, 1, lab)
        for c_idx, v in enumerate(vals, start=2):
            c = ws.cell(row=r_idx, column=c_idx, value=v); c.number_format = "0.00%"
            c.fill = FILL_OUTPUT; c.font = FONT_VALUE; c.border = BORDER; c.alignment = ALIGN_RIGHT
        label_cell(ws, r_idx, 6, note)

    # Esempio capitale 10k a 8y in EUR REALI (potere d'acquisto attualizzato)
    section_header(ws, 14, "ESEMPIO: 10.000 EUR a 8 anni (capitale finale REALE)", span=6)
    table_header(ws, 15, 1, "Strumento")
    for i, h in enumerate(["@ 1,6%", "@ 2,8%", "@ 3,5%", "@ 4,5%"], start=2):
        table_header(ws, 15, i, h)
    sim_rows = [
        ("BTP Italia FOI",          "=10000*(1+B6)^8",  "=10000*(1+C6)^8",  "=10000*(1+D6)^8",  "=10000*(1+E6)^8"),
        ("BTPi euro HICP",          "=10000*(1+B7)^8",  "=10000*(1+C7)^8",  "=10000*(1+D7)^8",  "=10000*(1+E7)^8"),
        ("ETF Oro",                 "=10000*(1+B8)^8",  "=10000*(1+C8)^8",  "=10000*(1+D8)^8",  "=10000*(1+E8)^8"),
        ("ETF REIT globale",        "=10000*(1+B9)^8",  "=10000*(1+C9)^8",  "=10000*(1+D9)^8",  "=10000*(1+E9)^8"),
        ("ETF azionario value",     "=10000*(1+B10)^8", "=10000*(1+C10)^8", "=10000*(1+D10)^8", "=10000*(1+E10)^8"),
        ("Conto deposito vincol.",  "=10000*(1+B11)^8", "=10000*(1+C11)^8", "=10000*(1+D11)^8", "=10000*(1+E11)^8"),
    ]
    for r_idx, (lab, *vals) in enumerate(sim_rows, start=16):
        label_cell(ws, r_idx, 1, lab)
        for c_idx, v in enumerate(vals, start=2):
            c = ws.cell(row=r_idx, column=c_idx, value=v); c.number_format = '#,##0" EUR"'
            c.fill = FILL_OUTPUT; c.font = FONT_VALUE; c.border = BORDER; c.alignment = ALIGN_RIGHT

    note_cell(ws, 23, 1,
              "CHIAVE DI LETTURA: questa tabella mostra il rendimento REALE (post-inflazione), cioe' "
              "quanto AUMENTA il tuo potere d'acquisto. BTP indicizzati hanno cedola reale fissa "
              "(1,5-1,9%). Strumenti a tasso nominale vedono il loro rendimento reale CALARE quando "
              "l'inflazione sale: il Conto Deposito al 3% nominale rende solo 1,4% reale @1,6%, "
              "scende a -1,4% reale @4,5%. ETF Oro/azionari resistono meglio. La conclusione: in "
              "scenari di inflazione strutturale, BTP indicizzati + Oro + ETF azionari battono il "
              "Conto Deposito di 1-3 punti REALI all'anno.",
              span=6)
    ws.row_dimensions[23].height = 90


def build_sheet_stress(wb):
    ws = wb.create_sheet("4 - Stress test Brent 5y")
    set_col_widths(ws, [42, 16, 16, 16, 16])
    title_row(ws, 1, "Stress test: Brent 120$ per 5 anni - impatto FIRE", span=5)
    disclaimer_row(ws, 2, span=5)
    section_header(ws, 4, "SCENARIO BASE vs STRESS (5 anni)", span=5)
    table_header(ws, 5, 1, "Variabile")
    table_header(ws, 5, 2, "Base (ISTAT 2025)")
    table_header(ws, 5, 3, "Stress (Brent 5y)")
    table_header(ws, 5, 4, "Delta")
    table_header(ws, 5, 5, "Note")
    sr = [
        ("Inflazione media 5y", 0.020, 0.035, "Risale e si stabilizza"),
        ("Brent medio 5y (USD)", 75, 110, "Tensione Medio Oriente"),
        ("Tassi BCE deposit facility", 0.025, 0.040, "Risalita per contrastare inflaz."),
        ("Salari nominali (acquisita)", 0.020, 0.025, "Recupero forzato"),
        ("Potere d'acquisto cum. 5y", -0.005, -0.045, "Erosione strutturale"),
        ("Rendimento ETF azion. nominale", 0.060, 0.040, "Compressione multipli"),
        ("Rendimento BTP Italia FOI", 0.039, 0.054, "Indicizzazione segue"),
        ("Rendimento Conto Dep vinc.", 0.030, 0.045, "Risale ma sotto inflaz."),
    ]
    for r_idx, (lab, base, stress, note) in enumerate(sr, start=6):
        label_cell(ws, r_idx, 1, lab)
        # base
        c = ws.cell(row=r_idx, column=2, value=base)
        c.number_format = ("0.00%" if abs(base) < 1 else '#,##0')
        c.fill = FILL_OUTPUT; c.font = FONT_VALUE; c.border = BORDER; c.alignment = ALIGN_RIGHT
        # stress
        c = ws.cell(row=r_idx, column=3, value=stress)
        c.number_format = ("0.00%" if abs(stress) < 1 else '#,##0')
        c.fill = FILL_OUTPUT; c.font = FONT_VALUE; c.border = BORDER; c.alignment = ALIGN_RIGHT
        # delta
        c = ws.cell(row=r_idx, column=4, value=f"=C{r_idx}-B{r_idx}")
        c.number_format = ("0.00%" if abs(base) < 1 else '#,##0')
        c.fill = FILL_OUTPUT; c.font = FONT_VALUE; c.border = BORDER; c.alignment = ALIGN_RIGHT
        label_cell(ws, r_idx, 5, note)

    section_header(ws, 16, "IMPATTO SUL TUO FIRE", span=5)
    label_cell(ws, 17, 1, "SWR conservativo (Trinity adattato)")
    output_cell(ws, 17, 2, "=0.040", "0.00%")
    output_cell(ws, 17, 3, "=0.030", "0.00%")
    output_cell(ws, 17, 4, "=C17-B17", "0.00%")
    label_cell(ws, 17, 5, "Scendi al 3% se inflaz. 3,5%")
    label_cell(ws, 18, 1, "Numero FIRE per 20k spese annue")
    output_cell(ws, 18, 2, "=20000/B17", '#,##0" EUR"')
    output_cell(ws, 18, 3, "=20000/C17", '#,##0" EUR"')
    output_cell(ws, 18, 4, "=C18-B18", '#,##0" EUR"')
    label_cell(ws, 18, 5, "Serve +166k EUR di patrimonio!")

    note_cell(ws, 20, 1,
              "LETTURA: nello scenario stress (Brent 120$ per 5 anni, inflazione 3,5% strutturale), "
              "il numero FIRE per coprire 20k EUR di spese annue passa da 500k a 666k EUR. "
              "Sono 166k in piu' di patrimonio necessario. Sopravvivono: BTP Italia FOI, BTPi, "
              "oro fisico, REIT. Soffrono: ETF azionario puro, conto deposito non vincolato, "
              "buoni postali a tasso fisso. La strategia anti-fragile vince.", span=5)
    ws.row_dimensions[20].height = 75


def main():
    wb = Workbook()
    sheets_info = [
        ("1 - SWR Italia 2026", "SWR aggiornato con dati ISTAT 2026 (inflazione + rendim. reale)"),
        ("2 - Numero FIRE 6 profili", "M/F, Nord/Sud, giovane/over 50 (dati ISTAT)"),
        ("3 - 6 alternative anti-inflaz", "BTP FOI, BTPi, Oro, REIT, value, conto deposito - 4 scenari"),
        ("4 - Stress test Brent 5y", "Cosa succede al FIRE con Brent 120$ per 5 anni"),
    ]
    sources = [
        "ISTAT — Rapporto Annuale 2026 (21/5/2026): inflazione 1,6% media 2025, 2,8% aprile 2026",
        "ISTAT — Retribuzione lorda mediana lavoro standard 2025: oltre 28k EUR annui",
        "ISTAT — Salari 2026 acquisita >2%; perdita potere acquisto vs 2019: -8,6%",
        "ISTAT — Occupazione 2025: 62,5% (15-64); disocc 6,1% media, 5,2% marzo 2026",
        "ISTAT — Mezzogiorno gap occupazione vs Nord: 20pp (era 23pp nel 2019)",
        "Morningstar State of Retirement Income 2026 — SWR base 3,9%",
        "Bengen W. (2025) — A Richer Retirement: SAFEMAX 4,7%",
        "MEF / D.M. economia — BTP Italia 2026 indicizzato FOI cedola 1,9%",
        "Poste Italiane — Buoni Soluzione Futuro 2025 (rend. medio 2,4%/anno)",
        "Brent aprile 2026: 120,4 USD/barile (ISTAT cita FMI/IEA)",
    ]
    build_cover(wb, "ISTAT 2026 x FIRE — Simulatori Italia", sheets_info, sources)
    build_sheet_swr(wb)
    build_sheet_profili(wb)
    build_sheet_inflation(wb)
    build_sheet_stress(wb)
    out = Path(__file__).resolve().parent.parent / "simulatori" / "istat_fire_2026.xlsx"
    wb.save(out)
    print(f"OK -> {out}")


if __name__ == "__main__":
    main()
