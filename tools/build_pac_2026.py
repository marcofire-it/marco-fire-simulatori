"""Genera pac_2026.xlsx — Simulatori PAC (Piano di Accumulo del Capitale)."""
from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _excel_helpers import (  # noqa: E402
    set_col_widths, title_row, disclaimer_row, section_header,
    label_cell, input_cell, output_cell, note_cell, table_header,
    build_cover,
)


def build_sheet_pac_base(wb):
    ws = wb.create_sheet("1 - PAC mensile")
    set_col_widths(ws, [42, 18, 4, 36, 18])
    title_row(ws, 1, "Simulatore PAC mensile")
    disclaimer_row(ws, 2)

    section_header(ws, 4, "INPUT")
    label_cell(ws, 5, 1, "Contributo mensile (€)"); input_cell(ws, 5, 2, 300, '#,##0" €"')
    label_cell(ws, 6, 1, "Anni di accumulo"); input_cell(ws, 6, 2, 25, "0")
    label_cell(ws, 7, 1, "Rendimento netto medio annuo (%)"); input_cell(ws, 7, 2, 0.05, "0.00%")
    label_cell(ws, 8, 1, "Capitale iniziale (€)"); input_cell(ws, 8, 2, 0, '#,##0" €"')

    note_cell(ws, 9, 1, "Rendimento netto stima ETF azionario globale lungo periodo: 5-7%. "
                       "Bilanciato 60/40: 4-5%. Obbligazionario: 1-3%.", span=2)
    ws.row_dimensions[9].height = 30

    section_header(ws, 11, "OUTPUT — capitale finale")
    # mesi totali
    label_cell(ws, 12, 1, "Mesi di accumulo")
    output_cell(ws, 12, 2, "=B6*12", "0")
    label_cell(ws, 13, 1, "Tasso mensile equivalente")
    output_cell(ws, 13, 2, "=(1+B7)^(1/12)-1", "0.0000%")
    label_cell(ws, 14, 1, "Totale versato")
    output_cell(ws, 14, 2, "=B5*B12+B8", '#,##0" €"')
    label_cell(ws, 15, 1, "Future Value PAC (formula annuity)")
    output_cell(ws, 15, 2, "=B5*(((1+B13)^B12-1)/B13)*(1+B13)", '#,##0" €"')
    label_cell(ws, 16, 1, "Future Value capitale iniziale composto")
    output_cell(ws, 16, 2, "=B8*(1+B7)^B6", '#,##0" €"')
    label_cell(ws, 17, 1, "Capitale finale LORDO totale")
    output_cell(ws, 17, 2, "=B15+B16", '#,##0" €"')
    label_cell(ws, 18, 1, "Plusvalenza (capitale - versato)")
    output_cell(ws, 18, 2, "=B17-B14", '#,##0" €"')
    label_cell(ws, 19, 1, "Aliquota capital gain (ETF azionario)")
    input_cell(ws, 19, 2, 0.26, "0.00%")
    label_cell(ws, 20, 1, "Capitale finale NETTO (dopo tasse plus)")
    output_cell(ws, 20, 2, "=B14+B18*(1-B19)", '#,##0" €"')

    note_cell(ws, 22, 1, "Nota: questa simulazione assume rendimento medio costante. "
                        "Nella realtà il PAC è soggetto a volatilità — il capitale finale può oscillare "
                        "±20-30% rispetto alla media in scenari storici.", span=2)
    ws.row_dimensions[22].height = 45


def build_sheet_pac_stepup(wb):
    ws = wb.create_sheet("2 - PAC con step-up")
    set_col_widths(ws, [42, 18, 4, 36, 18])
    title_row(ws, 1, "PAC con incremento annuo (step-up)")
    disclaimer_row(ws, 2)

    section_header(ws, 4, "INPUT")
    label_cell(ws, 5, 1, "Contributo mensile iniziale (€)"); input_cell(ws, 5, 2, 300, '#,##0" €"')
    label_cell(ws, 6, 1, "Incremento annuo del contributo (% es. inflazione)"); input_cell(ws, 6, 2, 0.025, "0.00%")
    label_cell(ws, 7, 1, "Anni di accumulo"); input_cell(ws, 7, 2, 25, "0")
    label_cell(ws, 8, 1, "Rendimento netto medio annuo (%)"); input_cell(ws, 8, 2, 0.05, "0.00%")

    note_cell(ws, 9, 1, "Lo step-up annuo aggiorna il contributo all'inflazione (mantiene il potere reale).",
              span=2)
    ws.row_dimensions[9].height = 22

    section_header(ws, 11, "OUTPUT")
    # Tabella anno per anno (max 30 anni)
    headers = ["Anno", "Contributo mensile", "Contributo annuo", "Capitale fine anno"]
    for i, h in enumerate(headers, start=1):
        table_header(ws, 13, i, h)

    # Anno 1: contributo iniziale (mensile)
    for y in range(1, 31):
        row = 13 + y
        label_cell(ws, row, 1, f"Anno {y}")
        # Contributo mensile = mensile iniziale × (1+step)^(y-1)
        output_cell(ws, row, 2, f"=B5*(1+B6)^({y-1})", '#,##0" €"')
        # Contributo annuo = mensile × 12
        output_cell(ws, row, 3, f"=B{row}_dummy", '#,##0" €"')
        # uso formula direttamente: =B{row}*12 — ma B{row} è la cell con contributo mensile, riferito direttamente
        ws.cell(row=row, column=3).value = f"=B{row}*12"
        # Capitale fine anno: ricorsivo (capitale precedente * (1+rend) + contributi annui composti)
        if y == 1:
            # FV di PAC anno 1: contributo mensile × ((1+r_m)^12 - 1)/r_m con r_m = (1+r)^(1/12)-1
            # Approssimazione semplificata: end-of-year sum + half-year growth
            ws.cell(row=row, column=4).value = f"=C{row}*((1+B8)^0.5)"
        else:
            ws.cell(row=row, column=4).value = f"=D{row-1}*(1+B8)+C{row}*((1+B8)^0.5)"
        # Stile output_cell per col D e C
        from _excel_helpers import FILL_OUTPUT, FONT_VALUE, BORDER, ALIGN_RIGHT
        for col in (2, 3, 4):
            c = ws.cell(row=row, column=col)
            c.fill = FILL_OUTPUT
            c.font = FONT_VALUE
            c.border = BORDER
            c.alignment = ALIGN_RIGHT
            c.number_format = '#,##0" €"'

    section_header(ws, 45, "RIEPILOGO (riferito ad 'Anni di accumulo' del campo input)")
    label_cell(ws, 46, 1, "Capitale finale (anno indicato in B7)")
    output_cell(ws, 46, 2, "=INDEX(D14:D43,B7)", '#,##0" €"')
    label_cell(ws, 47, 1, "Totale versato (somma contributi annui)")
    output_cell(ws, 47, 2, "=SUMPRODUCT((ROW(C14:C43)-13<=B7)*C14:C43)", '#,##0" €"')


def build_sheet_pac_vs_lumpsum(wb):
    ws = wb.create_sheet("3 - PAC vs lump sum")
    set_col_widths(ws, [42, 18, 4, 36, 18])
    title_row(ws, 1, "PAC vs investimento singolo (lump sum)")
    disclaimer_row(ws, 2)

    section_header(ws, 4, "INPUT")
    label_cell(ws, 5, 1, "Capitale totale da investire (€)"); input_cell(ws, 5, 2, 50000, '#,##0" €"')
    label_cell(ws, 6, 1, "Anni di orizzonte"); input_cell(ws, 6, 2, 15, "0")
    label_cell(ws, 7, 1, "Rendimento netto medio annuo (%)"); input_cell(ws, 7, 2, 0.06, "0.00%")
    label_cell(ws, 8, 1, "Durata spalmatura PAC (mesi)"); input_cell(ws, 8, 2, 12, "0")

    section_header(ws, 10, "OUTPUT")
    label_cell(ws, 11, 1, "PAC mensile equivalente (capitale / mesi)")
    output_cell(ws, 11, 2, "=B5/B8", '#,##0" €"')
    label_cell(ws, 12, 1, "Capitale finale LUMP SUM dopo X anni")
    output_cell(ws, 12, 2, "=B5*(1+B7)^B6", '#,##0" €"')
    label_cell(ws, 13, 1, "Capitale finale PAC (durata accumulo + investito)")
    # Semplifico: PAC accumula su B8 mesi, poi cresce per (B6 - B8/12) anni
    output_cell(ws, 13, 2,
                "=(B11*(((1+(1+B7)^(1/12)-1)^B8-1)/((1+B7)^(1/12)-1)))*(1+B7)^(B6-B8/12)",
                '#,##0" €"')
    label_cell(ws, 14, 1, "Differenza assoluta (LUMP - PAC)")
    output_cell(ws, 14, 2, "=B12-B13", '#,##0" €"')
    label_cell(ws, 15, 1, "Differenza in %")
    output_cell(ws, 15, 2, "=IFERROR((B12-B13)/B13, 0)", "0.00%")

    note_cell(ws, 17, 1, "Dati Vanguard 2023: lump sum batte DCA nel 68% dei casi a 12 mesi. "
                        "Outperformance media +2,2% (100% equity) / +1,8% (60/40). "
                        "Per dettaglio scenari crash/positive drift vedi simulatore PIC vs PAC dedicato.",
              span=2)
    ws.row_dimensions[17].height = 45


def build_sheet_pac_to_fire(wb):
    ws = wb.create_sheet("4 - PAC verso il FIRE")
    set_col_widths(ws, [42, 18, 4, 36, 18])
    title_row(ws, 1, "PAC e raggiungimento del numero FIRE")
    disclaimer_row(ws, 2)

    section_header(ws, 4, "INPUT")
    label_cell(ws, 5, 1, "Spese annue stimate post-FIRE (€)")
    input_cell(ws, 5, 2, 24000, '#,##0" €"')
    label_cell(ws, 6, 1, "Safe Withdrawal Rate (%)")
    input_cell(ws, 6, 2, 0.04, "0.00%")
    label_cell(ws, 7, 1, "Contributo mensile attuale (€)")
    input_cell(ws, 7, 2, 700, '#,##0" €"')
    label_cell(ws, 8, 1, "Capitale già accumulato (€)")
    input_cell(ws, 8, 2, 20000, '#,##0" €"')
    label_cell(ws, 9, 1, "Rendimento netto medio annuo (%)")
    input_cell(ws, 9, 2, 0.05, "0.00%")

    note_cell(ws, 10, 1, "SWR consigliato: 4% (Trinity Study) per 30 anni di pensione. "
                       "3-3,5% per FIRE precoce (40+ anni). 3,9% Morningstar 2026.", span=2)
    ws.row_dimensions[10].height = 30

    section_header(ws, 12, "OUTPUT")
    label_cell(ws, 13, 1, "Numero FIRE = spese / SWR")
    output_cell(ws, 13, 2, "=B5/B6", '#,##0" €"')
    label_cell(ws, 14, 1, "Tasso mensile equivalente")
    output_cell(ws, 14, 2, "=(1+B9)^(1/12)-1", "0.0000%")

    # Stima anni per arrivare al FIRE
    # FV(PAC) + FV(initial) = FIRE → risolvo per N mesi (formula chiusa logaritmo)
    label_cell(ws, 15, 1, "Stima anni per arrivare al FIRE")
    output_cell(ws, 15, 2,
                "=IFERROR(LN((B13*B14+B7)/(B8*B14+B7))/LN(1+B14)/12, NA())",
                "0.0")

    label_cell(ws, 17, 1, "Capitale finale stimato al traguardo")
    output_cell(ws, 17, 2, "=B13", '#,##0" €"')
    label_cell(ws, 18, 1, "Rendita mensile sostenibile (= spese / 12)")
    output_cell(ws, 18, 2, "=B5/12", '#,##0" €"')

    section_header(ws, 20, "Scenari — aumentare contributo accelera FIRE")
    table_header(ws, 21, 1, "Scenario")
    table_header(ws, 21, 2, "Contributo mensile")
    table_header(ws, 21, 3, "Anni stimati al FIRE")
    for i, mult in enumerate([0.75, 1.0, 1.25, 1.5, 2.0], start=22):
        label_cell(ws, i, 1, f"× {mult:.2f} del contributo attuale")
        output_cell(ws, i, 2, f"=B7*{mult}", '#,##0" €"')
        output_cell(ws, i, 3,
                    f"=IFERROR(LN((B13*B14+B7*{mult})/(B8*B14+B7*{mult}))/LN(1+B14)/12, NA())",
                    "0.0")


def main():
    wb = Workbook()
    sheets_info = [
        ("1 - PAC mensile", "Calcolatore PAC con contributo costante"),
        ("2 - PAC con step-up", "PAC con incremento annuo (inflazione)"),
        ("3 - PAC vs lump sum", "PAC spalmato vs investimento singolo"),
        ("4 - PAC verso il FIRE", "Anni per arrivare al numero FIRE col tuo PAC"),
    ]
    sources = [
        "Vanguard Research (2023) — Lump sum vs DCA",
        "Trinity Study + Morningstar 2026 (SWR 3,9-4%)",
        "Agenzia delle Entrate — Tassazione plusvalenze ETF 26% / 12,5% (titoli Stato white-list)",
        "Bengen 1994 + 2024 update — 4,7-5,25%",
    ]
    build_cover(wb, "Marco FIRE — Simulatori PAC 2026", sheets_info, sources)
    build_sheet_pac_base(wb)
    build_sheet_pac_stepup(wb)
    build_sheet_pac_vs_lumpsum(wb)
    build_sheet_pac_to_fire(wb)
    out = Path(__file__).resolve().parent.parent / "simulatori" / "pac_2026.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"saved: {out}")


if __name__ == "__main__":
    main()
