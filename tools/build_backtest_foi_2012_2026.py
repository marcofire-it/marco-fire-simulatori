"""build_backtest_foi_2012_2026.py — Back-test struttura BTP Italia "Sì" (2026) vs
BTP Italia classico, sul percorso reale dell'inflazione FOI ex-tabacchi 2012-2026.

OBIETTIVO: rendere RIPRODUCIBILE il claim del video post-emissione (blocco 09b):
  "+19€ floor senza watermark / -19€ cedola su capitale fermo = pari", su 1.000 €.

DIFFERENZA STRUTTURALE (fonti MEF / idealista / traderlink):
  - Classico: capitale RIVALUTATO dall'inflazione; la cedola reale fissa si applica al
    capitale rivalutato (cresce nel tempo). In deflazione, il recupero della deflazione
    pregressa avviene sulle cedole successive (meccanismo "watermark": l'inflazione nuova
    prima ripristina il picco precedente, poi torna a pagare).
  - Sì: capitale FERMO a 100 (rimborso a 100); la cedola fissa è sul nominale, l'inflazione
    del semestre è aggiunta sul nominale; NESSUN recupero della deflazione (floor per
    semestre, senza watermark): appena un semestre è positivo la variabile viene pagata.

I DUE EFFETTI (su capitale C, orizzonte H anni, tasso reale fisso r, tassazione tax):
  EFFETTO A (svantaggio Sì, "cedola su capitale fermo"):
     il classico incassa r anche sulla RIVALUTAZIONE del capitale; il Sì solo su 100.
     A_lordo = r * C * Σ_t (CI_t - 1)         (CI_t = indice FOI cumulato all'anno t)
     A_netto = A_lordo * (1 - tax)            -> entra col segno MENO nel Sì
  EFFETTO B (vantaggio Sì, "floor senza watermark"):
     il Sì non deve recuperare la deflazione pregressa -> "salva" la deflazione cumulata
     che il classico avrebbe dovuto recuperare prima di tornare a pagare la variabile.
     B_lordo = defl_recuperata * C            (defl_recuperata = % di deflazione recuperata)
     B_netto = B_lordo * (1 - tax)            -> entra col segno PIÙ nel Sì
  NETTO = B_netto - A_netto

ATTENZIONE GRANULARITÀ (punto chiave dell'audit):
  defl_recuperata dipende dalla risoluzione dei dati. A livello ANNUALE la deflazione FOI
  2012-2026 è quasi nulla (solo 2015 -0,2%, 2016 -0,1%, 2020 -0,2% = 0,5% totale) -> B≈+€4.
  Per arrivare a "pari" serve defl_recuperata ≈ 2,2% (effetto A netto), ottenibile solo
  con i CALI SEMESTRALI/MENSILI (più profondi della media annua). Quindi:
    - col dato ANNUALE: Sì ~ -15€/1000 (leggermente peggio del classico);
    - il "pari" del video regge SOLO se la deflazione semestrale recuperata ≈ 2,2%.
  Lo script lascia 'defl_recuperata' come PARAMETRO (vedi DEFL_RECUPERATA_*).

Uso:
  python build_backtest_foi_2012_2026.py
Output:
  - stampa i due effetti + netto (annuale e scenario "pari")
  - simulatori/backtest_foi_2012_2026.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _excel_helpers import (  # noqa: E402
    set_col_widths, title_row, disclaimer_row, section_header,
    label_cell, input_cell, output_cell, note_cell, table_header,
    build_cover, FILL_OUTPUT, FONT_VALUE, BORDER, ALIGN_RIGHT,
)
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # noqa: E402
from openpyxl.utils import get_column_letter as _gcl  # noqa: E402

# ---- Riquadro guida "per principianti" (a destra dei dati) ----
_G_TITLE_FILL = PatternFill("solid", fgColor="1E40AF")
_G_TITLE_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
_G_BODY_FILL = PatternFill("solid", fgColor="FFF9E6")
_G_BODY_FONT = Font(name="Calibri", size=11, color="333333")
_G_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
_G_ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_G_SIDE = Side(style="thin", color="C9A227")
_G_BORDER = Border(left=_G_SIDE, right=_G_SIDE, top=_G_SIDE, bottom=_G_SIDE)


def guide_box(ws, row, col, title, paragraphs, width=80):
    """Scrive un riquadro 'come leggere' in colonna `col` (singola, larga), a destra dei dati."""
    ws.column_dimensions[_gcl(col)].width = width
    t = ws.cell(row=row, column=col, value=title)
    t.fill = _G_TITLE_FILL; t.font = _G_TITLE_FONT; t.alignment = _G_ALIGN_C; t.border = _G_BORDER
    ws.row_dimensions[row].height = 24
    rr = row + 1
    chars_per_line = max(20, int(width * 1.05))
    for para in paragraphs:
        c = ws.cell(row=rr, column=col, value=para)
        c.fill = _G_BODY_FILL; c.font = _G_BODY_FONT; c.alignment = _G_ALIGN; c.border = _G_BORDER
        n = max(1, -(-len(para) // chars_per_line))     # ceil
        n = max(n, para.count("\n") + 1)
        ws.row_dimensions[rr].height = max(16, n * 16)
        rr += 1
    return rr

# ---- DATI: serie FOI ex-tabacchi, variazione media annua (ISTAT via rivaluta.it) -----
# 2012-2025 = 14 anni pieni ("ultimi 14 anni dal 2012"); 2026 parziale (YTD ~1,8%).
FOI_ANNUAL = [
    (2012, 0.030), (2013, 0.012), (2014, 0.002), (2015, -0.002), (2016, -0.001),
    (2017, 0.012), (2018, 0.011), (2019, 0.004), (2020, -0.002), (2021, 0.018),
    (2022, 0.080), (2023, 0.054), (2024, 0.009), (2025, 0.014),
]
FOI_2026_PARZIALE = 0.018

# ---- PARAMETRI back-test ----
CAPITALE = 1000.0
R_REAL = 0.016          # tasso reale fisso = 1,6% (tasso definitivo BTP Italia Sì giu-2026);
                        # uguale per entrambe le strutture, isola la differenza strutturale
TAX = 0.125
# defl_recuperata: % di deflazione che il classico deve recuperare (watermark) e il Sì no.
DEFL_RECUPERATA_ANNUALE = 0.005   # somma |deflazione annua| 2015+2016+2020 = 0,5% (lower bound)
# Stima SEMESTRALE realistica: profondità dei cali SOTTO IL PICCO precedente per un titolo
# emesso nel 2012. La corsa 2012-2014 (+3,2%) cuscina i cali successivi:
#   - 2015-2016: trough gen-2016 = 99,7 (base 2015=100) vs picco ~100,3 (2014) -> ~0,6%
#   - 2020 (COVID): dip ~0,5% sotto il picco 2019
#   tot recuperato ~1,1% (NON 2,2%: la deflazione italiana e' stata mite).
DEFL_RECUPERATA_SEMESTRALE = 0.011


def _index_path(series):
    """CI cumulato (base fine-2011 = 1,0)."""
    ci, out = 1.0, []
    for anno, var in series:
        ci *= (1 + var)
        out.append((anno, var, ci))
    return out


def compute():
    path = _index_path(FOI_ANNUAL)
    sum_ci_minus_1 = sum(ci - 1 for _, _, ci in path)     # Σ(CI_t - 1)
    cumulata = path[-1][2] - 1                              # +X% a fine 2025
    defl_annuale = sum(-var for _, var, _ in path if var < 0)  # somma |deflazione annua|

    a_lordo = R_REAL * CAPITALE * sum_ci_minus_1
    a_netto = a_lordo * (1 - TAX)
    b_netto_annuale = DEFL_RECUPERATA_ANNUALE * CAPITALE * (1 - TAX)
    b_netto_sem = DEFL_RECUPERATA_SEMESTRALE * CAPITALE * (1 - TAX)
    netto_annuale = b_netto_annuale - a_netto
    netto_sem = b_netto_sem - a_netto       # stima realistica semestrale
    # defl che serve per "pari" (B_netto = A_netto):
    defl_pari = a_netto / (CAPITALE * (1 - TAX))

    return {
        "path": path, "sum_ci_minus_1": sum_ci_minus_1, "cumulata": cumulata,
        "defl_annuale": defl_annuale, "a_lordo": a_lordo, "a_netto": a_netto,
        "b_netto_annuale": b_netto_annuale, "netto_annuale": netto_annuale,
        "b_netto_sem": b_netto_sem, "netto_sem": netto_sem,
        "defl_pari": defl_pari,
    }


def _print(r):
    print("=" * 64)
    print("BACK-TEST FOI 2012-2025 — struttura Sì vs classico (su 1.000 €)")
    print("=" * 64)
    print(f"  FOI cumulata 2012-2025      : +{r['cumulata']*100:.1f}%")
    print(f"  Σ(CI_t - 1) [14 anni]       : {r['sum_ci_minus_1']:.3f}")
    print(f"  Deflazione ANNUALE recuper. : {r['defl_annuale']*100:.1f}%")
    print("-" * 64)
    print(f"  EFFETTO A (cedola su rivalut.)  r={R_REAL*100:.1f}%:")
    print(f"     lordo = {r['a_lordo']:.2f} €   netto = -{r['a_netto']:.2f} €")
    print(f"  EFFETTO B (floor senza watermark):")
    print(f"     annuale (lower bound, defl 0,5%)  netto = +{r['b_netto_annuale']:.2f} €")
    print(f"     semestrale stima (defl ~1,1%)     netto = +{r['b_netto_sem']:.2f} €")
    print("-" * 64)
    print(f"  NETTO Sì vs classico — annuale     = {r['netto_annuale']:.2f} €")
    print(f"  NETTO Sì vs classico — semestrale  = {r['netto_sem']:.2f} €  <-- stima realistica")
    print("-" * 64)
    print(f"  Per 'PARI' serve deflazione recuperata ≈ {r['defl_pari']*100:.2f}%")
    print(f"  (reale ~1,1% -> il Sì resta ~{-r['netto_sem']:.0f}€ SOTTO: 'pari' del video ottimistico)")
    print("=" * 64)


# --------------------------------------------------------------------------- EXCEL
def build_excel(r):
    wb = Workbook()
    sheets = [
        ("1 - Serie FOI 2012-2026", "Inflazione FOI ex-tabacchi annua + indice cumulato (ISTAT)"),
        ("2 - Backtest Sì vs classico", "I due effetti (cedola su rivalutato / floor) e il netto"),
        ("3 - Sensibilità", "Netto al variare di tasso fisso e deflazione recuperata"),
    ]
    sources = [
        "ISTAT — indice FOI ex-tabacchi, variazioni medie annue 2012-2026 (via rivaluta.it / I.Stat)",
        "MEF Dip. Tesoro — BTP Italia Sì 2026: capitale rimborsato a 100, cedola su nominale, floor deflazione SENZA recupero (watermark)",
        "idealista/news, traderlink — confronto struttura Sì vs BTP Italia classico (rivalutazione capitale + recupero deflazione)",
        "Back-test = ricalcolo proprio: i due effetti su 1.000 € e percorso FOI reale (vedi sheet 2)",
    ]
    build_cover(wb, "Back-test BTP Italia Sì vs classico — FOI 2012-2026", sheets, sources)
    guide_box(wb["0 - Indice"], 4, 4, "📖 IN PAROLE SEMPLICI", [
        "A cosa serve questo file: rispondere a una domanda semplice — il NUOVO BTP Italia "
        "'Sì' del 2026 conviene quanto il VECCHIO BTP Italia, oppure è peggiorato?",
        "Come lo scopriamo: simuliamo 1.000 € investiti e usiamo l'inflazione italiana VERA "
        "degli ultimi 14 anni (2012-2026), non numeri inventati.",
        "Le schede in breve:  • '1 - Serie FOI' = l'inflazione reale, anno per anno.  "
        "• '2 - Backtest' = il confronto vero e proprio tra i due titoli.  "
        "• '3 - Sensibilità' = come cambia il risultato se cambi le ipotesi.",
        "Il risultato in due righe: il nuovo Sì rende circa 11 € in meno su 1.000 € rispetto "
        "al vecchio, in questo scenario storico. NON è una fregatura, ma è un filo meno "
        "generoso. Tutti i dettagli nella scheda 2.",
        "Ricorda: celle GIALLE = le puoi modificare; celle VERDI = si calcolano da sole.",
    ])

    _sheet_serie(wb, r)
    _sheet_backtest(wb, r)
    _sheet_sensi(wb, r)

    out = Path(__file__).resolve().parent.parent / "simulatori" / "backtest_foi_2012_2026.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def _sheet_serie(wb, r):
    ws = wb.create_sheet("1 - Serie FOI 2012-2026")
    set_col_widths(ws, [12, 18, 18, 22])
    title_row(ws, 1, "Serie FOI ex-tabacchi 2012-2026 (ISTAT)", span=4)
    disclaimer_row(ws, 2, span=4)
    section_header(ws, 4, "INFLAZIONE FOI ANNUA + INDICE CUMULATO", span=4)
    for col, h in enumerate(["Anno", "Var. FOI %", "Indice cumulato (CI)", "CI - 1 (rivalut.)"], start=1):
        table_header(ws, 5, col, h)
    row = 6
    for anno, var, ci in r["path"]:
        label_cell(ws, row, 1, anno)
        c = ws.cell(row=row, column=2, value=var); c.number_format = "0.0%"
        c.fill = FILL_OUTPUT; c.font = FONT_VALUE; c.border = BORDER; c.alignment = ALIGN_RIGHT
        output_cell(ws, row, 3, ci, "0.000")
        output_cell(ws, row, 4, ci - 1, "0.0%")
        row += 1
    label_cell(ws, row, 1, "2026 (parz.)")
    c = ws.cell(row=row, column=2, value=FOI_2026_PARZIALE); c.number_format = "0.0%"
    c.fill = FILL_OUTPUT; c.font = FONT_VALUE; c.border = BORDER; c.alignment = ALIGN_RIGHT
    row += 2
    label_cell(ws, row, 1, "FOI cumulata 2012-2025")
    output_cell(ws, row, 2, r["cumulata"], "0.0%")
    note_cell(ws, row + 2, 1,
              "L'indice CI parte da 1,000 a fine 2011. La deflazione ANNUA è quasi assente "
              "(solo 2015 -0,2%, 2016 -0,1%, 2020 -0,2%): per questo l'effetto B (floor) calcolato "
              "sui dati annuali è piccolo. I cali SEMESTRALI/mensili sono più profondi.", span=4)

    guide_box(ws, 4, 6, "📖 COME LEGGERE QUESTA SCHEDA", [
        "Cos'è il FOI? È l'indice ISTAT che misura l'inflazione: di quanto aumentano i "
        "prezzi in Italia ogni anno. Il BTP Italia usa proprio questo indice per proteggere "
        "i tuoi soldi dall'inflazione.",
        "Colonna 'Var. FOI %' = quanto sono saliti i prezzi in quell'anno. Esempio: 2022 "
        "+8% = nel 2022 il carrello della spesa è diventato più caro dell'8%. Quando il "
        "numero è negativo (2015, 2016, 2020) i prezzi sono leggermente scesi (deflazione).",
        "Colonna 'Indice cumulato (CI)' = quanto sono saliti i prezzi DALL'INIZIO del 2012, "
        "tutto sommato. Parte da 1,000 e arriva a ~1,266: vuol dire che dal 2012 a oggi i "
        "prezzi sono cresciuti del +26,6% in totale.",
        "Perché ti serve: con un BTP Italia, più i prezzi salgono più alta è la cedola che "
        "incassi. Questa scheda mostra cos'è successo DAVVERO ai prezzi, così i conti delle "
        "altre schede partono da numeri reali, non inventati.",
    ])


def _sheet_backtest(wb, r):
    ws = wb.create_sheet("2 - Backtest Sì vs classico")
    set_col_widths(ws, [40, 16, 16, 18, 16, 16])
    title_row(ws, 1, "Back-test struttura Sì vs classico — su 1.000 €", span=6)
    disclaimer_row(ws, 2, span=6)

    section_header(ws, 4, "PARAMETRI (celle gialle = modificabili)", span=6)
    label_cell(ws, 5, 1, "Capitale nominale (€)");      input_cell(ws, 5, 2, CAPITALE, '#,##0" €"')
    label_cell(ws, 6, 1, "Tasso reale fisso annuo r");  input_cell(ws, 6, 2, R_REAL, "0.0%")
    label_cell(ws, 7, 1, "Tassazione titoli stato");    input_cell(ws, 7, 2, TAX, "0.0%")

    section_header(ws, 9, "EFFETTO A — cedola su capitale rivalutato (svantaggio Sì)", span=6)
    for col, h in enumerate(["Anno", "Var FOI %", "Indice CI", "Cedola classico (r·CI)",
                             "Cedola Sì (r·100)", "Extra classico"], start=1):
        table_header(ws, 10, col, h)
    first = 11
    for i, (anno, var, ci) in enumerate(r["path"]):
        rr = first + i
        label_cell(ws, rr, 1, anno)
        c = ws.cell(row=rr, column=2, value=var); c.number_format = "0.0%"
        c.fill = FILL_OUTPUT; c.font = FONT_VALUE; c.border = BORDER; c.alignment = ALIGN_RIGHT
        prev = "1" if i == 0 else f"C{rr-1}"
        output_cell(ws, rr, 3, f"={prev}*(1+B{rr})", "0.000")
        output_cell(ws, rr, 4, f"=$B$6*C{rr}*$B$5", '#,##0.00" €"')
        output_cell(ws, rr, 5, "=$B$6*$B$5", '#,##0.00" €"')
        output_cell(ws, rr, 6, f"=D{rr}-E{rr}", '#,##0.00" €"')
    last = first + len(r["path"]) - 1
    rsum = last + 1
    label_cell(ws, rsum, 1, "Somma extra classico (LORDO)")
    output_cell(ws, rsum, 6, f"=SUM(F{first}:F{last})", '#,##0.00" €"')
    ra = rsum + 1
    label_cell(ws, ra, 1, "EFFETTO A — netto (Sì incassa MENO)")
    output_cell(ws, ra, 2, f"=-F{rsum}*(1-$B$7)", '#,##0.00" €"')

    rb_h = ra + 2
    section_header(ws, rb_h, "EFFETTO B — floor senza watermark (vantaggio Sì)", span=6)
    rb_in = rb_h + 1
    label_cell(ws, rb_in, 1, "Deflazione cumulata RECUPERATA (input)")
    input_cell(ws, rb_in, 2, DEFL_RECUPERATA_SEMESTRALE, "0.0%")
    note_cell(ws, rb_in, 3,
              "Default 1,1% = stima SEMESTRALE realistica (cali sotto-picco: trough gen-2016 99,7 "
              "vs picco 2014 ~0,6% + dip 2020 ~0,5%). Annuale puro = 0,5%. Per 'pari' servirebbe "
              "~2,2% (non supportato dalla deflazione italiana, mite).", span=4)
    rb = rb_in + 1
    label_cell(ws, rb, 1, "EFFETTO B — netto (Sì risparmia)")
    output_cell(ws, rb, 2, f"=B{rb_in}*$B$5*(1-$B$7)", '#,##0.00" €"')

    rn_h = rb + 2
    section_header(ws, rn_h, "RISULTATO NETTO Sì vs classico", span=6)
    rn = rn_h + 1
    label_cell(ws, rn, 1, "NETTO (B - A): >0 Sì meglio, <0 Sì peggio")
    output_cell(ws, rn, 2, f"=B{rb}+B{ra}", '#,##0.00" €"')
    label_cell(ws, rn + 1, 1, "Deflazione recuperata per il 'PARI'")
    output_cell(ws, rn + 1, 2, f"=-B{ra}/($B$5*(1-$B$7))", "0.00%")

    note_cell(ws, rn + 3, 1,
              f"CONCLUSIONE: stima realistica (deflazione recuperata ~1,1% semestrale) -> il Sì esce "
              f"~{-r['netto_sem']:.0f}€/1.000 SOTTO il classico, NON 'pari'. Effetto A (-{r['a_netto']:.0f}€, "
              f"cedola su capitale rivalutato) e' robusto; l'effetto B (floor) e' piccolo perche' la deflazione "
              f"italiana 2012-2026 e' stata mite e la corsa 2012-2014 cuscina i cali. Il 'pari' del video "
              f"richiederebbe ~{r['defl_pari']*100:.1f}% di deflazione recuperata (ottimistico). RACCOMANDAZIONE: "
              "ammorbidire il claim a 'molto simili, Sì leggermente sotto salvo deflazione marcata'. "
              "NB: l'effetto B esatto dipende dalla regola di recupero del vecchio prospetto MEF. NON e' consulenza.", span=6)

    guide_box(ws, 4, 8, "📖 COME LEGGERE QUESTA SCHEDA", [
        "La domanda: il NUOVO BTP Italia 'Sì' (2026) è una fregatura rispetto al VECCHIO "
        "BTP Italia? Mettiamo 1.000 € e calcoliamo la differenza usando l'inflazione VERA "
        "del 2012-2026.",
        "Ci sono due differenze tra i due titoli, due 'effetti' che tirano in direzioni opposte:",
        "▶ EFFETTO A (≈ −20 €, svantaggio del Sì): nel vecchio la cedola fissa si calcolava "
        "su un capitale che cresceva con l'inflazione; nel Sì si calcola sempre su 100. Quindi "
        "col Sì incassi un po' MENO di cedola nel tempo.",
        "▶ EFFETTO B (≈ +10 €, vantaggio del Sì): se i prezzi scendono (deflazione), il Sì non "
        "ti obbliga a 'recuperare' il calo prima di tornare a pagarti l'inflazione; il vecchio sì. "
        "Su questi 14 anni questo vantaggio vale circa +10 €.",
        "▶ NETTO ≈ −11 € : sommando A e B, il Sì rende circa 11 € in meno su 1.000 € rispetto al "
        "vecchio, in questo scenario storico. Tradotto: NON è una fregatura, ma neanche identico — "
        "è LEGGERMENTE meno generoso.",
        "Celle GIALLE = puoi modificarle (capitale, tasso, tassazione): i numeri verdi si "
        "ricalcolano da soli. Prova a cambiare il capitale e guarda come scala tutto.",
        "In breve: il 'pari' di cui si parla in giro vale solo se ci fosse stata MOLTA deflazione "
        "(~2,4%). In Italia ce n'è stata poca, quindi il Sì resta un filo sotto il vecchio.",
    ])


def _sheet_sensi(wb, r):
    ws = wb.create_sheet("3 - Sensibilità")
    set_col_widths(ws, [34, 16, 16, 16, 16])
    title_row(ws, 1, "Sensibilità del NETTO (Sì vs classico, €/1.000)", span=5)
    disclaimer_row(ws, 2, span=5)
    section_header(ws, 4, "NETTO al variare di: tasso fisso r (righe) × deflazione recuperata (col)", span=5)

    defl_cols = [0.005, 0.011, 0.020, 0.0235]
    r_rows = [0.010, 0.015, 0.016, 0.020]
    table_header(ws, 5, 1, "r reale \\ deflaz. recup.")
    for j, d in enumerate(defl_cols, start=2):
        table_header(ws, 5, j, f"{d*100:.1f}%")
    for i, rr in enumerate(r_rows, start=6):
        label_cell(ws, i, 1, f"{rr*100:.1f}%")
        a_net = rr * CAPITALE * r["sum_ci_minus_1"] * (1 - TAX)
        for j, d in enumerate(defl_cols, start=2):
            net = d * CAPITALE * (1 - TAX) - a_net
            c = ws.cell(row=i, column=j, value=net); c.number_format = '#,##0.0" €"'
            c.fill = FILL_OUTPUT; c.font = FONT_VALUE; c.border = BORDER; c.alignment = ALIGN_RIGHT

    note_cell(ws, 11, 1,
              f"I valori vicini a 0 = 'pari'. A r={R_REAL*100:.1f}% (Sì) il 'pari' richiede deflazione "
              f"recuperata ≈ {r['defl_pari']*100:.1f}% (ultima colonna). Con la stima realistica 1,1% "
              f"(2a colonna) il Sì resta ~{-r['netto_sem']:.0f}€ sotto; con la deflazione solo annuale 0,5% "
              f"(1a colonna) ancora di più. Σ(CI-1) = {r['sum_ci_minus_1']:.3f} (FOI reale 2012-2025).", span=5)

    guide_box(ws, 4, 7, "📖 COME LEGGERE QUESTA SCHEDA", [
        "Cos'è: una tabella 'e se...?'. Mostra come cambia il risultato (€ guadagnati o persi "
        "col Sì rispetto al vecchio, su 1.000 €) al variare di due ipotesi.",
        "RIGHE = il tasso fisso reale del titolo. Più è alto, più pesa l'Effetto A, cioè più "
        "il Sì è penalizzato. La riga 1,6% è quella del BTP Italia Sì vero.",
        "COLONNE = quanta deflazione il vecchio titolo deve 'recuperare'. Più ce n'è, più il "
        "Sì è avvantaggiato (Effetto B). La 2a colonna (1,1%) è la stima realistica per il 2012-2026.",
        "Come si legge un numero: NEGATIVO = il Sì rende MENO del vecchio; vicino a ZERO = vanno "
        "PARI; POSITIVO = il Sì rende di PIÙ. Cerca i valori vicini allo zero per capire quando "
        "si pareggiano.",
        "Conclusione: incrocia la riga 1,6% con la colonna 1,1% (i valori reali) → il Sì è ~11 € "
        "sotto. Si arriva a 'pari' solo nell'ultima colonna (deflazione ~2,4%), che però NON si "
        "è verificata in Italia.",
    ])


def main():
    r = compute()
    _print(r)
    out = build_excel(r)
    print(f"\nExcel -> {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
