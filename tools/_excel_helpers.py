"""Helper condivisi per la generazione di simulatori Excel.

Usato da tutti gli script build_<slug>_2026.py per produrre xlsx con stile uniforme
(celle gialle = input, verdi = output, header blu, disclaimer in alto).
"""
from __future__ import annotations
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ====== Stili condivisi ======
FILL_INPUT = PatternFill("solid", fgColor="FFF59D")
FILL_OUTPUT = PatternFill("solid", fgColor="C8E6C9")
FILL_HEADER = PatternFill("solid", fgColor="1E40AF")
FILL_NOTE = PatternFill("solid", fgColor="FAFAFA")
FILL_DISCLAIMER = PatternFill("solid", fgColor="FEE2E2")
FILL_TABLE_HEADER = PatternFill("solid", fgColor="DBEAFE")

FONT_HEADER = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
FONT_LABEL = Font(name="Calibri", size=11, bold=True)
FONT_VALUE = Font(name="Calibri", size=11)
FONT_NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
FONT_TITLE = Font(name="Calibri", size=16, bold=True, color="1E40AF")
FONT_DISCLAIMER = Font(name="Calibri", size=10, italic=True, color="991B1B")
FONT_SECTION_H1 = Font(name="Calibri", size=14, bold=True)

THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

DISCLAIMER_TXT = (
    "Strumento informativo / divulgativo. NON è consulenza finanziaria. "
    "I valori sono indicativi al 2026 e vanno verificati col tuo consulente."
)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_row(ws, row, text, span=4):
    ws.cell(row=row, column=1, value=text).font = FONT_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 25


def disclaimer_row(ws, row, span=4):
    c = ws.cell(row=row, column=1, value=DISCLAIMER_TXT)
    c.font = FONT_DISCLAIMER
    c.fill = FILL_DISCLAIMER
    c.alignment = ALIGN_LEFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 32


def section_header(ws, row, text, span=4):
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_HEADER
    c.fill = FILL_HEADER
    c.alignment = ALIGN_CENTER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def label_cell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = FONT_LABEL
    c.alignment = ALIGN_LEFT
    c.border = BORDER


def input_cell(ws, row, col, value, fmt="0"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = FILL_INPUT
    c.font = FONT_VALUE
    c.border = BORDER
    c.alignment = ALIGN_RIGHT
    c.number_format = fmt
    return c


def output_cell(ws, row, col, formula, fmt="0"):
    c = ws.cell(row=row, column=col, value=formula)
    c.fill = FILL_OUTPUT
    c.font = FONT_VALUE
    c.border = BORDER
    c.alignment = ALIGN_RIGHT
    c.number_format = fmt
    return c


def table_header(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = FONT_LABEL
    c.fill = FILL_TABLE_HEADER
    c.alignment = ALIGN_CENTER
    c.border = BORDER
    return c


def note_cell(ws, row, col, text, span=1):
    c = ws.cell(row=row, column=col, value=text)
    c.font = FONT_NOTE
    c.fill = FILL_NOTE
    c.alignment = ALIGN_LEFT
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    # Stima altezza riga in base a lunghezza testo + larghezza colonne span
    # Larghezza totale span in "character units" Excel
    from openpyxl.utils import get_column_letter
    total_width = 0
    for c_idx in range(col, col + span):
        dim = ws.column_dimensions.get(get_column_letter(c_idx))
        total_width += (dim.width if dim and dim.width else 10)
    chars_per_line = max(20, int(total_width * 0.95))  # 95% per padding
    n_lines = max(1, (len(text) // chars_per_line) + (1 if len(text) % chars_per_line else 0))
    # Conta anche newline espliciti
    n_lines = max(n_lines, text.count('\n') + 1)
    # Altezza riga = n_lines * 15pt (line height standard FONT_NOTE 9pt)
    needed_height = n_lines * 15
    # Se la riga ha gia' un'altezza maggiore, non ridurla
    current_h = ws.row_dimensions[row].height or 15
    ws.row_dimensions[row].height = max(current_h, needed_height)
    return c


def section_h1(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = FONT_SECTION_H1
    return c


def build_cover(wb, title_text, sheets_info, sources, github_url="github.com/marcofire-it/marco-fire-simulatori"):
    """Costruisce la sheet 'Indice' con titolo, disclaimer, indice, come si usa, fonti.

    sheets_info: lista di tuple (sheet_name, description).
    sources: lista di stringhe (fonti dati).
    """
    ws = wb.active
    ws.title = "0 - Indice"
    set_col_widths(ws, [60, 30])

    title_row(ws, 1, title_text, span=2)
    ws.row_dimensions[1].height = 32

    disclaimer_row(ws, 2, span=2)

    section_h1(ws, 4, 1, "Indice")
    for i, (name, desc) in enumerate(sheets_info, start=5):
        ws.cell(row=i, column=1, value=name).font = FONT_LABEL
        ws.cell(row=i, column=2, value=desc).font = FONT_VALUE
        ws.row_dimensions[i].height = 22

    base = 5 + len(sheets_info) + 1
    section_h1(ws, base, 1, "Come usare")
    ws.cell(row=base + 1, column=1, value="1. Le celle GIALLE sono input — modificale coi tuoi numeri.").font = FONT_VALUE
    ws.cell(row=base + 2, column=1, value="2. Le celle VERDI sono output — si aggiornano automaticamente.").font = FONT_VALUE
    ws.cell(row=base + 3, column=1, value="3. Tutti i numeri di default sono al 2026 da fonti pubbliche.").font = FONT_VALUE
    ws.cell(row=base + 4, column=1, value="4. Niente codice nascosto: l'Excel è solo formule modificabili.").font = FONT_VALUE

    base2 = base + 6
    section_h1(ws, base2, 1, "Fonti dati")
    for i, s in enumerate(sources, start=base2 + 1):
        # Fonte parametrica = FORMULA: scrivila as-is (deve includere gia' il
        # bullet "• " nel testo); stringa semplice = bullet aggiunto qui.
        val = s if isinstance(s, str) and s.startswith("=") else f"• {s}"
        ws.cell(row=i, column=1, value=val).font = FONT_VALUE

    base3 = base2 + 1 + len(sources) + 1
    ws.cell(row=base3, column=1, value=f"Aggiornamenti: {github_url}").font = FONT_NOTE
