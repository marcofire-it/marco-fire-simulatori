"""Build btp_si_compare_2026.xlsx - confronto BTP Italia Si vs altri BTP retail.

Fogli:
 0 - Indice
 1 - Parametri (input centralizzato — un solo posto da cambiare il 12 giu)
 2 - BTP Italia Si (5y, fisso + FOI semestrale, premio 0,6%)
 3 - BTP Italia classico (capitale rivalutato + cedola reale)
 4 - BTP Valore (6y step-up + premio 0,8% + YTM a prezzo MOT per chi compra oggi)
 5 - BTP Futura (12y step-up + premio PIL storico + YTM a prezzo MOT per chi compra oggi)
 6 - BTP nominali (2y/5y/10y curva fissa)
 7 - Matrice comparativa (6 strumenti x 4 scenari inflazione)
 8 - Break-even multi-strumento (soglia inflazione per Si vs ognuno)
 9 - 4 profili FIRE (allocazione)
10 - Calcolatore personale

Output (default): e:/sviluppo/marco-fire-simulatori-staging/simulatori/btp_si_compare_2026.xlsx
(repo STAGING private; promozione a public solo via tools/release_excel.py al go-live).
Override: python tools/build_btp_si_compare_2026.py --out <path>
"""
from __future__ import annotations
import sys
from pathlib import Path

THIS_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(THIS_DIR))

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import ColorScaleRule

from _excel_helpers import (
    set_col_widths, title_row, disclaimer_row, section_header,
    label_cell, input_cell, output_cell, table_header, note_cell, build_cover,
)


def _style_axes(ch, y_fmt="0.00%", x_fmt=None, plot_x=0.12, plot_y=0.14, plot_w=0.74, plot_h=0.68):
    """Configura assi: tick labels visibili + gridlines chiare + chart area con margini extra
    per evitare label appiccicate ai titoli/bordi.
    plot_x/y/w/h: layout del PLOT AREA INNER (zona dati) rispetto al box del chart.
    """
    # Visibility
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    # Tick labels position
    ch.x_axis.tickLblPos = "low"
    ch.y_axis.tickLblPos = "nextTo"
    # Number format
    if y_fmt:
        ch.y_axis.number_format = y_fmt
    if x_fmt:
        ch.x_axis.number_format = x_fmt
    # Gridlines: grigio chiaro (opacità ridotta via colore)
    gp = GraphicalProperties()
    gp.line = LineProperties(solidFill="D0D0D0", w=4500)
    gl_y = ChartLines()
    gl_y.spPr = gp
    ch.y_axis.majorGridlines = gl_y
    # No gridlines verticali (rumore visivo)
    ch.x_axis.majorGridlines = None
    # Plot area layout: inner = solo zona dati
    ch.plot_area.layout = Layout(manualLayout=ManualLayout(
        x=plot_x, y=plot_y, w=plot_w, h=plot_h,
        xMode="edge", yMode="edge",
        layoutTarget="inner",
    ))
    # SPOSTA il Y title verticale al bordo SINISTRO del chart (lontano dai tick labels Y)
    if ch.y_axis.title is not None:
        ch.y_axis.title.layout = Layout(manualLayout=ManualLayout(
            x=0.005, y=0.30,   # quasi attaccato al bordo sx, centrato verticalmente sull'area
            xMode="edge", yMode="edge",
        ))
        ch.y_axis.title.overlay = False
    # SPOSTA il X title in basso, centrato, lontano dai tick labels X
    if ch.x_axis.title is not None:
        ch.x_axis.title.layout = Layout(manualLayout=ManualLayout(
            x=0.40, y=0.93,
            xMode="edge", yMode="edge",
        ))
        ch.x_axis.title.overlay = False
    # CHART TITLE: non sovrapporre al plot area (forza posizione in alto)
    if ch.title is not None:
        ch.title.overlay = False
    # LEGENDA: forza fuori plot area (lato destro)
    if ch.legend is not None:
        ch.legend.overlay = False
        ch.legend.position = "r"


def add_bar_chart(ws, title, data_range, cat_range, anchor, y_title="", x_title="", style=10, bar_dir="col", y_fmt="0.00%"):
    """Aggiunge un bar chart standard.

    data_range / cat_range: tuple (min_col, min_row, max_col, max_row)
    """
    ch = BarChart()
    ch.type = bar_dir
    ch.style = style
    ch.title = title
    if y_title:
        ch.y_axis.title = y_title
    if x_title:
        ch.x_axis.title = x_title
    data = Reference(ws, min_col=data_range[0], min_row=data_range[1],
                     max_col=data_range[2], max_row=data_range[3])
    cats = Reference(ws, min_col=cat_range[0], min_row=cat_range[1],
                     max_col=cat_range[2], max_row=cat_range[3])
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ch.width = 18
    ch.height = 11
    # Plot area con margini ampi per evitare sovrapposizione Y title <-> tick labels
    if bar_dir == "bar":
        # Bar orizzontale: categorie sulla Y, ancora più spazio sinistro
        _style_axes(ch, y_fmt=y_fmt, plot_x=0.32, plot_y=0.16, plot_w=0.60, plot_h=0.66)
    else:
        _style_axes(ch, y_fmt=y_fmt, plot_x=0.20, plot_y=0.16, plot_w=0.72, plot_h=0.66)
    ws.add_chart(ch, anchor)
    return ch


def add_line_chart(ws, title, data_range, cat_range, anchor, y_title="", x_title="", style=12, y_fmt="0.00%"):
    ch = LineChart()
    ch.style = style
    ch.title = title
    if y_title:
        ch.y_axis.title = y_title
    if x_title:
        ch.x_axis.title = x_title
    data = Reference(ws, min_col=data_range[0], min_row=data_range[1],
                     max_col=data_range[2], max_row=data_range[3])
    cats = Reference(ws, min_col=cat_range[0], min_row=cat_range[1],
                     max_col=cat_range[2], max_row=cat_range[3])
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ch.width = 22
    ch.height = 13
    _style_axes(ch, y_fmt=y_fmt, plot_x=0.18, plot_y=0.16, plot_w=0.74, plot_h=0.66)
    ws.add_chart(ch, anchor)
    return ch


def build_params(wb):
    """Foglio 1 - Parametri centralizzati. UNICO posto da modificare il 12 giu."""
    ws = wb.create_sheet("1 - Parametri")
    set_col_widths(ws, [50, 18, 14, 50])
    SPAN = 4  # larghezza UNIFORME bande/sezioni = max colonna tabelle del foglio (col D)

    title_row(ws, 1, "PARAMETRI CENTRALIZZATI (modifica QUI, tutto il resto si aggiorna)", span=SPAN)
    disclaimer_row(ws, 2, span=SPAN)

    section_header(ws, 4, "BTP Italia Si - emissione 15-19 giugno 2026", span=SPAN)
    rows = [
        ("Tasso fisso reale annuo (possibile rialzo a chiusura 19/06)", 0.012, "0.00%", "Tasso minimo garantito annunciato dal MEF in apertura emissione. Puo' essere rivisto AL RIALZO (mai al ribasso) alla chiusura del collocamento, ven 19/06/2026 ore 13. Se rivisto, aggiorna SOLO questa cella e tutte le simulazioni si ricalcolano."),
        ("Durata (anni)",                                  5,     "0",     "Fissa: 5 anni"),
        ("Premio fedelta' finale",                         0.006, "0.00%", "Confermato MEF: 0,6% del capitale (solo se mantenuto a scadenza)"),
        ("Tassazione titoli di stato",                     0.125, "0.0%",  "12,5% su cedole + premio finale"),
    ]
    for i, (lbl, val, fmt, note) in enumerate(rows, start=5):
        label_cell(ws, i, 1, lbl)
        input_cell(ws, i, 2, val, fmt=fmt)
        note_cell(ws, i, 4, note)

    section_header(ws, 10, "BTP Italia classico in essere (riferimenti secondario MOT)", span=SPAN)
    rows = [
        ("BTP Italia mar 2028 - cedola reale annua",  0.020, "0.00%", "MEF IT0005532723, scadenza 14/03/2028: tasso reale 2,0% sul prospetto (rendimento per chi compra a 100)."),
        ("BTP Italia giu 2030 - cedola reale annua",  0.016, "0.00%", "MEF IT0005497000, scadenza 28/06/2030: tasso reale 1,6% sul prospetto"),
    ]
    for i, (lbl, val, fmt, note) in enumerate(rows, start=11):
        label_cell(ws, i, 1, lbl)
        input_cell(ws, i, 2, val, fmt=fmt)
        note_cell(ws, i, 4, note)

    section_header(ws, 14, "BTP Valore (7a emissione marzo 2026, ISIN IT0005696338, secondario)", span=SPAN)
    rows = [
        ("Cedola step-up anni 1-2",  0.0260, "0.00%", "MEF: 2,60% definitivo (cedole trimestrali)"),
        ("Cedola step-up anni 3-4",  0.0320, "0.00%", "MEF: 3,20% definitivo (il 2,80 era il minimo garantito pre-collocamento)"),
        ("Cedola step-up anni 5-6",  0.0380, "0.00%", "MEF: 3,80% (definitivo)"),
        ("Premio fedelta'",          0.008,  "0.00%", "MEF: 0,8% del capitale a scadenza (solo sottoscrittori in collocamento)"),
        ("Durata (anni)",            6,      "0",     "Scadenza 10 mar 2032"),
    ]
    for i, (lbl, val, fmt, note) in enumerate(rows, start=15):
        label_cell(ws, i, 1, lbl)
        input_cell(ws, i, 2, val, fmt=fmt)
        note_cell(ws, i, 4, note)

    section_header(ws, 21, "BTP Futura 4a emissione nov 2021 (sospeso, ISIN IT0005466351 sul MOT)", span=SPAN)
    rows = [
        ("Cedola step-up anni 1-4",  0.0075, "0.00%", "MEF 4a emissione nov 2021, step DEFINITIVI: 0,75% (NON 0,35 = 2a emissione)"),
        ("Cedola step-up anni 5-8",  0.0135, "0.00%", "MEF definitivo: 1,35% (anni 5-8)"),
        ("Cedola step-up anni 9-12", 0.0170, "0.00%", "MEF definitivo: 1,70% (anni 9-12)"),
        ("Premio fedelta' min (PIL)", 0.004, "0.00%", "Floor 0,4% se PIL nominale basso (solo sottoscrittori originali)"),
        ("Premio fedelta' max (PIL)", 0.030, "0.00%", "Cap 3,0% se PIL nominale alto (solo sottoscrittori originali)"),
        ("Durata (anni)",            12,     "0",     "4a emissione (4+4+4 step-up), scadenza 16/11/2033"),
    ]
    for i, (lbl, val, fmt, note) in enumerate(rows, start=22):
        label_cell(ws, i, 1, lbl)
        input_cell(ws, i, 2, val, fmt=fmt)
        note_cell(ws, i, 4, note)

    section_header(ws, 29, "BTP nominali - curva 2026 (asta MEF)", span=SPAN)
    rows = [
        ("BTP 2y - rendimento lordo (feb 2028)",   0.0220, "0.00%", "MEF asta: 2,20%"),
        ("BTP 3y - rendimento lordo (mar 2029)",   0.0240, "0.00%", "MEF asta: 2,40%"),
        ("BTP 5y - rendimento lordo (stima)",      0.0290, "0.00%", "Stima interpolata 2y-7y curva 2026"),
        ("BTP 7y - rendimento lordo (mar 2033)",   0.0315, "0.00%", "MEF asta: 3,15%"),
        ("BTP 10y - rendimento lordo (feb 2036)",  0.0345, "0.00%", "MEF asta: 3,45%"),
    ]
    for i, (lbl, val, fmt, note) in enumerate(rows, start=30):
        label_cell(ws, i, 1, lbl)
        input_cell(ws, i, 2, val, fmt=fmt)
        note_cell(ws, i, 4, note)

    section_header(ws, 36, "Scenari inflazione FOI media 5 anni (per simulazioni)", span=SPAN)
    rows = [
        ("Scenario 1: Disinflazione",   0.010, "0.0%", "BCE stretta riuscita - inflazione 1,0%"),
        ("Scenario 2: Base BCE",        0.020, "0.0%", "Target 2,0% medio termine"),
        ("Scenario 3: Persistente",     0.030, "0.0%", "Inflazione struturale 3,0%"),
        ("Scenario 4: Stress geopol.",  0.045, "0.0%", "Brent 120 USD persistente - 4,5%"),
    ]
    for i, (lbl, val, fmt, note) in enumerate(rows, start=37):
        label_cell(ws, i, 1, lbl)
        input_cell(ws, i, 2, val, fmt=fmt)
        note_cell(ws, i, 4, note)

    section_header(ws, 42, "Setup simulazioni", span=SPAN)
    rows = [
        ("Capitale investito (EUR)",  10000, "#,##0", "Importo scalabile per tutte le simulazioni"),
    ]
    for i, (lbl, val, fmt, note) in enumerate(rows, start=43):
        label_cell(ws, i, 1, lbl)
        input_cell(ws, i, 2, val, fmt=fmt)
        note_cell(ws, i, 4, note)

    # ==== Sezione AGGIUNTIVA: prezzi secondario + YTM reale effettivo ====
    section_header(ws, 45, "BTP Italia classico - PREZZI MOT (YTM reale effettivo per chi compra al secondario)", span=SPAN)

    label_cell(ws, 46, 1, "Prezzo secondario BTP Italia mar 2028")
    input_cell(ws, 46, 2, 101.97, fmt="0.00")
    label_cell(ws, 46, 3, "Anni residui")
    input_cell(ws, 46, 4, 1.8, fmt="0.0")
    note_cell(ws, 47, 1, "Quotazione ufficiale MOT 09/06/2026: 101,97 (100 = par). Aggiorna col valore corrente quando cambia il mercato.", span=SPAN)

    label_cell(ws, 48, 1, "Prezzo secondario BTP Italia giu 2030")
    input_cell(ws, 48, 2, 101.92, fmt="0.00")
    label_cell(ws, 48, 3, "Anni residui")
    input_cell(ws, 48, 4, 4, fmt="0.0")  # giu 2030 - giu 2026 = 4 anni residui
    note_cell(ws, 49, 1, "Quotazione ufficiale MOT 09/06/2026: 101,92. Aggiorna col valore corrente quando cambia il mercato.", span=SPAN)

    # YTM reale effettivo (bond equivalent yield approssimato)
    # YTM = (cedola + (100 - prezzo)/N) / ((100 + prezzo)/2)
    label_cell(ws, 50, 1, "YTM reale mar 2028 (per chi compra al MOT oggi)")
    output_cell(ws, 50, 2, "=(B11*100+(100-B46)/D46)/((100+B46)/2)", fmt="0.00%")
    note_cell(ws, 51, 1, "Formula bond equiv. yield. Cedola 2,0% sul prospetto, ma a prezzo 101,97 con 1,8y residui → YTM reale ~0,9%.", span=SPAN)

    label_cell(ws, 52, 1, "YTM reale giu 2030 (per chi compra al MOT oggi)")
    output_cell(ws, 52, 2, "=(B12*100+(100-B48)/D48)/((100+B48)/2)", fmt="0.00%")
    note_cell(ws, 53, 1, "Cedola 1,6% sul prospetto, a prezzo 101,92 con 4y residui → YTM reale ~1,1%.", span=SPAN)

    section_header(ws, 55, "YTM reale medio classico per il confronto matrice (media ponderata 2 ISIN)", span=SPAN)
    label_cell(ws, 56, 1, "YTM reale effettivo (media mar 2028 + giu 2030)")
    output_cell(ws, 56, 2, "=(B50+B52)/2", fmt="0.00%")
    note_cell(ws, 57, 1, "Usato come riferimento nelle matrici e break-even. Se vuoi cambiare la pesatura, modifica la formula.", span=SPAN)

    # ==== YTM reale NETTO (i numeri citati nel video) ====
    # NB: NON spostare B5-B48/D46/D48 (usati da sync_btp_params.py EXCEL_CELL_MAP).
    section_header(ws, 58, "YTM reale NETTO (citati nel video)", span=SPAN)
    label_cell(ws, 59, 1, "YTM reale NETTO mar 2028")
    output_cell(ws, 59, 2, "=B50*(1-B8)", fmt="0.00%")
    label_cell(ws, 60, 1, "YTM reale NETTO giu 2030")
    output_cell(ws, 60, 2, "=B52*(1-B8)", fmt="0.00%")
    label_cell(ws, 61, 1, "YTM reale NETTO medio (2 ISIN)")
    output_cell(ws, 61, 2, "=B56*(1-B8)", fmt="0.00%")

    # ==== Prezzi MOT BTP Valore / BTP Futura (per YTM di chi compra OGGI al secondario) ====
    # NB: righe 63-73 mappate in sync_btp_params.py EXCEL_CELL_MAP
    #     (B64/D64 Valore, B66/D66 Futura, B70:B73 anagrafica ISIN). NON spostare.
    section_header(ws, 63, "BTP Valore / BTP Futura - PREZZI MOT (per YTM di chi compra oggi)", span=SPAN)
    label_cell(ws, 64, 1, "Prezzo MOT BTP Valore mar 2026")
    input_cell(ws, 64, 2, 98.65, fmt="0.00")
    label_cell(ws, 64, 3, "Anni residui")
    input_cell(ws, 64, 4, 5.75, fmt="0.00")
    note_cell(ws, 65, 1, "Quotazione ufficiale MOT 09/06/2026: 98,65 (SOTTO la pari). Scadenza 10/03/2032. Il premio fedelta' 0,8% NON spetta a chi compra sul MOT.", span=SPAN)
    label_cell(ws, 66, 1, "Prezzo MOT BTP Futura nov 2033")
    input_cell(ws, 66, 2, 86.15, fmt="0.00")
    label_cell(ws, 66, 3, "Anni residui")
    input_cell(ws, 66, 4, 7.4, fmt="0.00")
    note_cell(ws, 67, 1, "Quotazione ufficiale MOT 09/06/2026: 86,15 (forte sconto sotto par → YTM lordo ~3,65%). Scadenza 16/11/2033. Il doppio premio PIL NON spetta a chi compra sul MOT.", span=SPAN)

    # ==== Anagrafica ISIN (sync da JSON master) ====
    section_header(ws, 69, "Anagrafica ISIN (Borsa Italiana MOT)", span=SPAN)
    isin_rows = [
        (70, "BTP Italia mar 2028", "IT0005532723", "Scadenza 14/03/2028"),
        (71, "BTP Italia giu 2030", "IT0005497000", "Scadenza 28/06/2030"),
        (72, "BTP Valore mar 2026", "IT0005696338", "7a emissione, scadenza 10/03/2032"),
        (73, "BTP Futura nov 2033", "IT0005466351", "4a emissione nov 2021, scadenza 16/11/2033"),
    ]
    for r, lbl, isin, note in isin_rows:
        label_cell(ws, r, 1, lbl)
        input_cell(ws, r, 2, isin, fmt="@")
        note_cell(ws, r, 4, note)

    return ws


def build_btp_si(wb):
    """Foglio 2 - BTP Italia Si simulazione 4 scenari."""
    ws = wb.create_sheet("2 - BTP Italia Si")
    set_col_widths(ws, [42, 16, 16, 16, 16, 16])
    SPAN = 6  # larghezza UNIFORME bande/sezioni = max colonna tabelle del foglio (col F, tabella IRR)

    title_row(ws, 1, "BTP Italia Si - rendimento netto in 4 scenari inflazione FOI", span=SPAN)
    disclaimer_row(ws, 2, span=SPAN)

    section_header(ws, 4, "Input (da foglio Parametri)", span=SPAN)
    label_cell(ws, 5, 1, "Tasso fisso reale annuo")
    output_cell(ws, 5, 2, "='1 - Parametri'!B5", fmt="0.00%")
    label_cell(ws, 6, 1, "Durata (anni)")
    output_cell(ws, 6, 2, "='1 - Parametri'!B6", fmt="0")
    label_cell(ws, 7, 1, "Premio fedelta' finale")
    output_cell(ws, 7, 2, "='1 - Parametri'!B7", fmt="0.00%")
    label_cell(ws, 8, 1, "Tassazione")
    output_cell(ws, 8, 2, "='1 - Parametri'!B8", fmt="0.0%")
    label_cell(ws, 9, 1, "Capitale investito (EUR)")
    output_cell(ws, 9, 2, "='1 - Parametri'!B43", fmt="#,##0")

    section_header(ws, 11, "Cedola annua lorda = fisso + inflazione FOI media", span=SPAN)
    table_header(ws, 12, 1, "Scenario inflazione")
    table_header(ws, 12, 2, "Infl. FOI media")
    table_header(ws, 12, 3, "Cedola lorda annua")
    table_header(ws, 12, 4, "Cedola netta annua")
    table_header(ws, 12, 5, "Importo netto su cap.")
    scen_rows = [
        (13, "1 - Disinflazione 1,0%",  "='1 - Parametri'!B37"),
        (14, "2 - Base BCE 2,0%",       "='1 - Parametri'!B38"),
        (15, "3 - Persistente 3,0%",    "='1 - Parametri'!B39"),
        (16, "4 - Stress geopol. 4,5%", "='1 - Parametri'!B40"),
    ]
    for r, lbl, ref in scen_rows:
        label_cell(ws, r, 1, lbl)
        output_cell(ws, r, 2, ref, fmt="0.0%")
        output_cell(ws, r, 3, f"=$B$5+B{r}", fmt="0.00%")
        output_cell(ws, r, 4, f"=C{r}*(1-$B$8)", fmt="0.00%")
        output_cell(ws, r, 5, f"=$B$9*D{r}", fmt="#,##0")

    section_header(ws, 18, "Rendimento totale finale 5 anni (netto, con premio 0,6%)", span=SPAN)
    table_header(ws, 19, 1, "Scenario")
    table_header(ws, 19, 2, "Cedole nette tot 5y")
    table_header(ws, 19, 3, "Premio netto fedelta'")
    table_header(ws, 19, 4, "Tot. netto incassato")
    table_header(ws, 19, 5, "IRR nom netto annuo (lineare)")
    table_header(ws, 19, 6, "IRR reale netto")
    for r, lbl, _ in scen_rows:
        rr = r + 7  # 20, 21, 22, 23
        label_cell(ws, rr, 1, lbl)
        output_cell(ws, rr, 2, f"=E{r}*$B$6", fmt="#,##0")
        output_cell(ws, rr, 3, f"=$B$9*$B$7*(1-$B$8)", fmt="#,##0")
        output_cell(ws, rr, 4, f"=B{rr}+C{rr}", fmt="#,##0")
        # IRR nominale netto LINEARE (no compounding): (fisso + premio/durata + FOI scenario)*(1-tasse)
        # FOI scenario in B{r} (righe 13-16 della tabella cedole). ROUND a 6 dp per tie deterministico col foglio 7.
        output_cell(ws, rr, 5, f"=ROUND(($B$5+$B$7/$B$6+B{r})*(1-$B$8),6)", fmt="0.00%")
        # IRR reale netto = nominale netto - FOI scenario
        output_cell(ws, rr, 6, f"=E{rr}-B{r}", fmt="0.00%")

    section_header(ws, 25, "NOTA: BTP Italia Si NON rivaluta il capitale (resta a 100)", span=SPAN)
    note_cell(ws, 26, 1, "Differenza chiave vs BTP Italia classico: qui l'inflazione "
                          "alimenta solo la cedola, il capitale a scadenza torna nominale.", span=SPAN)

    # Bar chart: IRR netto annuo per scenario (col E, rows 19-23, header in row 19)
    add_bar_chart(
        ws,
        title="BTP Italia Sì - IRR netto annuo per scenario",
        data_range=(5, 19, 5, 23),  # E19:E23 (col E = IRR; row 19 = header)
        cat_range=(1, 20, 1, 23),   # A20:A23 (4 scenari)
        anchor="H4",  # tabelle fino a col F: lascia G vuota di distacco
        y_title="IRR netto",
    )
    # Heatmap IRR netto annuo: rosso (basso) -> verde (alto)
    color_rule_si = ColorScaleRule(
        start_type='min', start_color='F87171',
        mid_type='percentile', mid_value=50, mid_color='FEF3C7',
        end_type='max', end_color='86EFAC',
    )
    ws.conditional_formatting.add('E20:E23', color_rule_si)
    return ws


def build_btp_classico(wb):
    """Foglio 3 - BTP Italia classico (capitale rivalutato)."""
    ws = wb.create_sheet("3 - BTP Italia classico")
    set_col_widths(ws, [42, 16, 16, 16, 16])
    SPAN = 5  # larghezza UNIFORME bande/sezioni = max colonna tabelle del foglio (col E)

    title_row(ws, 1, "BTP Italia classico - capitale rivalutato per inflazione FOI", span=SPAN)
    disclaimer_row(ws, 2, span=SPAN)

    section_header(ws, 4, "Input (BTP Italia giugno 2030 come riferimento)", span=SPAN)
    label_cell(ws, 5, 1, "Tasso reale annuo (giu 2030)")
    output_cell(ws, 5, 2, "='1 - Parametri'!B12", fmt="0.00%")
    label_cell(ws, 6, 1, "Durata residua simulata (anni)")
    input_cell(ws, 6, 2, 5, fmt="0")
    label_cell(ws, 7, 1, "Tassazione")
    output_cell(ws, 7, 2, "='1 - Parametri'!B8", fmt="0.0%")
    label_cell(ws, 8, 1, "Capitale investito (EUR)")
    output_cell(ws, 8, 2, "='1 - Parametri'!B43", fmt="#,##0")

    section_header(ws, 10, "Meccanismo classico: cedola reale + capitale rivalutato", span=SPAN)
    table_header(ws, 11, 1, "Scenario inflazione")
    table_header(ws, 11, 2, "FOI cumulata 5y")
    table_header(ws, 11, 3, "Capitale rivalutato")
    table_header(ws, 11, 4, "Cedole nette 5y (su cap. rivalut.)")
    table_header(ws, 11, 5, "Tot. netto 5y (cedole + rivalut.)")
    scen_rows = [
        (12, "1 - Disinflazione 1,0%",  "='1 - Parametri'!B37"),
        (13, "2 - Base BCE 2,0%",       "='1 - Parametri'!B38"),
        (14, "3 - Persistente 3,0%",    "='1 - Parametri'!B39"),
        (15, "4 - Stress geopol. 4,5%", "='1 - Parametri'!B40"),
    ]
    for r, lbl, ref in scen_rows:
        label_cell(ws, r, 1, lbl)
        output_cell(ws, r, 2, ref, fmt="0.0%")
        # Capitale rivalutato = cap * (1+infl)^anni (importo lordo cumulato)
        output_cell(ws, r, 3, f"=$B$8*(1+B{r})^$B$6", fmt="#,##0")
        # Cedole nette: tasso_reale * cap * sum((1+infl)^k per k=1..n) * (1-tasse)
        # Approx. continua: tasso * cap * ((1+infl)^n - 1)/infl * (1-tasse)
        output_cell(ws, r, 4, f"=$B$5*$B$8*((1+B{r})^$B$6-1)/B{r}*(1-$B$7)", fmt="#,##0")
        # Totale netto = cedole nette + rivalutazione capitale NETTA (anche la rivalut. e' tassata 12,5%)
        output_cell(ws, r, 5, f"=(C{r}-$B$8)*(1-$B$7)+D{r}", fmt="#,##0")

    section_header(ws, 17, "IRR netto annuo a confronto", span=SPAN)
    table_header(ws, 18, 1, "Scenario")
    table_header(ws, 18, 2, "Tot. netto 5y")
    table_header(ws, 18, 3, "IRR netto annuo")
    for r, lbl, _ in scen_rows:
        rr = r + 7  # 19, 20, 21, 22
        label_cell(ws, rr, 1, lbl)
        output_cell(ws, rr, 2, f"=E{r}", fmt="#,##0")
        output_cell(ws, rr, 3, f"=((E{r}+$B$8)/$B$8)^(1/$B$6)-1", fmt="0.00%")

    section_header(ws, 24, "NOTA: vantaggio classico vs Si quando inflazione e' alta (capitale +)", span=SPAN)
    note_cell(ws, 25, 1, "Classico: l'inflazione cresce IL CAPITALE (effetto compounding). "
                          "Si: l'inflazione cresce SOLO la cedola annua. "
                          "Trade-off: il classico costa di piu' sul secondario (no premio se compri usato).", span=SPAN)

    # Bar chart: IRR netto annuo per scenario (col C = IRR, rows 18-22)
    add_bar_chart(
        ws,
        title="BTP Italia classico - IRR netto annuo per scenario",
        data_range=(3, 18, 3, 22),  # C18:C22 (header in C18)
        cat_range=(1, 19, 1, 22),   # A19:A22
        anchor="G4",
        y_title="IRR netto",
    )
    # Heatmap IRR netto
    color_rule_cl = ColorScaleRule(
        start_type='min', start_color='F87171',
        mid_type='percentile', mid_value=50, mid_color='FEF3C7',
        end_type='max', end_color='86EFAC',
    )
    ws.conditional_formatting.add('C19:C22', color_rule_cl)
    # Anche col E (Tot. netto 5y) per il primo blocco
    ws.conditional_formatting.add('E12:E15', color_rule_cl)
    return ws


def build_btp_valore(wb):
    """Foglio 4 - BTP Valore step-up."""
    ws = wb.create_sheet("4 - BTP Valore")
    set_col_widths(ws, [42, 16, 16, 16, 16])
    SPAN = 5  # larghezza UNIFORME bande/sezioni = max colonna tabelle del foglio (col E)

    title_row(ws, 1, "BTP Valore (marzo 2026) - step-up 2+2+2 cedole trimestrali", span=SPAN)
    disclaimer_row(ws, 2, span=SPAN)

    section_header(ws, 4, "Input", span=SPAN)
    label_cell(ws, 5, 1, "Cedola anni 1-2")
    output_cell(ws, 5, 2, "='1 - Parametri'!B15", fmt="0.00%")
    label_cell(ws, 6, 1, "Cedola anni 3-4")
    output_cell(ws, 6, 2, "='1 - Parametri'!B16", fmt="0.00%")
    label_cell(ws, 7, 1, "Cedola anni 5-6")
    output_cell(ws, 7, 2, "='1 - Parametri'!B17", fmt="0.00%")
    label_cell(ws, 8, 1, "Premio fedelta'")
    output_cell(ws, 8, 2, "='1 - Parametri'!B18", fmt="0.00%")
    label_cell(ws, 9, 1, "Durata (anni)")
    output_cell(ws, 9, 2, "='1 - Parametri'!B19", fmt="0")
    label_cell(ws, 10, 1, "Tassazione")
    output_cell(ws, 10, 2, "='1 - Parametri'!B8", fmt="0.0%")
    label_cell(ws, 11, 1, "Capitale investito")
    output_cell(ws, 11, 2, "='1 - Parametri'!B43", fmt="#,##0")

    section_header(ws, 13, "Cedole nette per fase step-up", span=SPAN)
    table_header(ws, 14, 1, "Fase")
    table_header(ws, 14, 2, "Cedola lorda annua")
    table_header(ws, 14, 3, "Cedola netta annua")
    table_header(ws, 14, 4, "Importo netto x anno")
    table_header(ws, 14, 5, "Importo netto x fase (2y)")
    fasi = [
        (15, "Anni 1-2 (2,60%)",  "B5"),
        (16, "Anni 3-4 (3,20%)",  "B6"),
        (17, "Anni 5-6 (3,80%)",  "B7"),
    ]
    for r, lbl, src in fasi:
        label_cell(ws, r, 1, lbl)
        output_cell(ws, r, 2, f"={src}", fmt="0.00%")
        output_cell(ws, r, 3, f"={src}*(1-$B$10)", fmt="0.000%")
        output_cell(ws, r, 4, f"=$B$11*C{r}", fmt="#,##0")
        output_cell(ws, r, 5, f"=D{r}*2", fmt="#,##0")

    section_header(ws, 19, "Rendimento totale netto 6 anni", span=SPAN)
    label_cell(ws, 20, 1, "Tot. cedole nette 6y")
    output_cell(ws, 20, 2, "=E15+E16+E17", fmt="#,##0")
    label_cell(ws, 21, 1, "Premio fedelta' netto")
    output_cell(ws, 21, 2, "=B11*B8*(1-B10)", fmt="#,##0")
    label_cell(ws, 22, 1, "Tot. netto 6y")
    output_cell(ws, 22, 2, "=B20+B21", fmt="#,##0")
    label_cell(ws, 23, 1, "IRR nom netto annuo (lineare)")
    # Lineare: (media cedole lorde step-up + premio/durata)*(1-tasse) -> ~2,92% con default
    # NB: vale per chi ha sottoscritto IN COLLOCAMENTO (paga 100, incassa il premio 0,8%).
    output_cell(ws, 23, 2, "=ROUND(((B5+B6+B7)/3+B8/B9)*(1-B10),6)", fmt="0.00%")

    section_header(ws, 25, "NOTA: il BTP Valore NON e' indicizzato all'inflazione", span=SPAN)
    note_cell(ws, 26, 1, "Cedole step-up FISSE in nominale. In inflazione alta perde potere d'acquisto reale. "
                          "Vantaggio: cedola trimestrale (vs semestrale altri BTP) per pianificazione cassa.", span=SPAN)

    # ==== Comprato OGGI sul MOT: YTM a prezzo di mercato (sotto la pari) ====
    section_header(ws, 28, "Comprato OGGI sul MOT - YTM a prezzo 98,65 (sotto la pari)", span=SPAN)
    label_cell(ws, 29, 1, "Prezzo MOT (uff. 09/06/2026)")
    output_cell(ws, 29, 2, "='1 - Parametri'!B64", fmt="0.00")
    label_cell(ws, 30, 1, "Anni residui (scad. 10/03/2032)")
    output_cell(ws, 30, 2, "='1 - Parametri'!D64", fmt="0.00")
    label_cell(ws, 31, 1, "Cedola media residua (step pesati sugli anni rimasti)")
    # Fasi ancorate alla scadenza: ultimi 2y = step3, 2y prima = step2, il resto = step1
    output_cell(ws, 31, 2, "=(B5*MAX(0,B30-4)+B6*MIN(2,MAX(0,B30-2))+B7*MIN(2,B30))/B30", fmt="0.00%")
    label_cell(ws, 32, 1, "YTM lordo (formula bond equivalent yield)")
    # BEY: (cedola_media + (100-P)/N) / ((100+P)/2) — stessa convenzione del foglio Parametri B50/B52
    output_cell(ws, 32, 2, "=(B31*100+(100-B29)/B30)/((100+B29)/2)", fmt="0.00%")
    label_cell(ws, 33, 1, "YTM netto (tassazione 12,5%)")
    output_cell(ws, 33, 2, "=B32*(1-B10)", fmt="0.00%")
    note_cell(ws, 34, 1, "Comprato OGGI SOTTO la pari (98,65): il mercato sconta tassi saliti dopo il collocamento, "
                          "quindi il YTM lordo (~3,48%) e' SUPERIORE alla media delle cedole (~3,23% sugli anni residui). "
                          "ATTENZIONE: il premio fedelta' 0,8% NON spetta a chi compra sul MOT (solo sottoscrittori in collocamento), "
                          "percio' il YTM qui sopra NON lo include.", span=SPAN)

    section_header(ws, 36, "IRR reale netto per scenario (da YTM MOT)", span=SPAN)
    table_header(ws, 37, 1, "Scenario inflazione")
    table_header(ws, 37, 2, "Infl. FOI media")
    table_header(ws, 37, 3, "IRR reale netto (YTM netto - FOI)")
    mot_scen = [
        (38, "1 - Disinflazione 1,0%",  "='1 - Parametri'!B37"),
        (39, "2 - Base BCE 2,0%",       "='1 - Parametri'!B38"),
        (40, "3 - Persistente 3,0%",    "='1 - Parametri'!B39"),
        (41, "4 - Stress geopol. 4,5%", "='1 - Parametri'!B40"),
    ]
    for r, lbl, ref in mot_scen:
        label_cell(ws, r, 1, lbl)
        output_cell(ws, r, 2, ref, fmt="0.0%")
        # Cedola nominale fissa: reale LINEARE = ytm_netto - FOI (convenzione foglio 7)
        output_cell(ws, r, 3, f"=$B$33-B{r}", fmt="0.00%")
    color_rule_val_mot = ColorScaleRule(
        start_type='min', start_color='F87171',
        mid_type='percentile', mid_value=50, mid_color='FEF3C7',
        end_type='max', end_color='86EFAC',
    )
    ws.conditional_formatting.add('C38:C41', color_rule_val_mot)

    # Bar chart: cedole lorde step-up (col B, rows 14-17)
    add_bar_chart(
        ws,
        title="BTP Valore - cedole step-up 2+2+2",
        data_range=(2, 14, 2, 17),  # B14:B17 (header in B14, 3 fasi)
        cat_range=(1, 15, 1, 17),   # A15:A17 (3 fasi)
        anchor="G4",
        y_title="Cedola annua lorda",
    )
    return ws


def build_btp_futura(wb):
    """Foglio 5 - BTP Futura (ipotesi premio PIL)."""
    ws = wb.create_sheet("5 - BTP Futura")
    set_col_widths(ws, [42, 16, 16, 16, 16, 16])
    SPAN = 6  # larghezza UNIFORME bande/sezioni = max colonna tabelle del foglio (col F, tabella PIL)

    title_row(ws, 1, "BTP Futura - sospeso, ricostruzione storica 4a emissione 12y", span=SPAN)
    disclaimer_row(ws, 2, span=SPAN)

    section_header(ws, 4, "Input", span=SPAN)
    label_cell(ws, 5, 1, "Cedola anni 1-4")
    output_cell(ws, 5, 2, "='1 - Parametri'!B22", fmt="0.00%")
    label_cell(ws, 6, 1, "Cedola anni 5-8")
    output_cell(ws, 6, 2, "='1 - Parametri'!B23", fmt="0.00%")
    label_cell(ws, 7, 1, "Cedola anni 9-12")
    output_cell(ws, 7, 2, "='1 - Parametri'!B24", fmt="0.00%")
    label_cell(ws, 8, 1, "Premio min (PIL basso)")
    output_cell(ws, 8, 2, "='1 - Parametri'!B25", fmt="0.00%")
    label_cell(ws, 9, 1, "Premio max (PIL alto)")
    output_cell(ws, 9, 2, "='1 - Parametri'!B26", fmt="0.00%")
    label_cell(ws, 10, 1, "Durata (anni)")
    output_cell(ws, 10, 2, "='1 - Parametri'!B27", fmt="0")
    label_cell(ws, 11, 1, "Tassazione")
    output_cell(ws, 11, 2, "='1 - Parametri'!B8", fmt="0.0%")
    label_cell(ws, 12, 1, "Capitale")
    output_cell(ws, 12, 2, "='1 - Parametri'!B43", fmt="#,##0")

    section_header(ws, 14, "Scenari di PIL nominale e premio fedelta'", span=SPAN)
    table_header(ws, 15, 1, "Scenario PIL nominale medio")
    table_header(ws, 15, 2, "PIL nom. medio")
    table_header(ws, 15, 3, "Premio fed. atteso")
    table_header(ws, 15, 4, "Cedole nette 12y")
    table_header(ws, 15, 5, "Tot. netto 12y")
    table_header(ws, 15, 6, "IRR nom netto annuo (lineare)")
    pil_scen = [
        (16, "Bassa (PIL 1,5%)",  0.015),
        (17, "Base (PIL 3,0%)",   0.030),
        (18, "Alta (PIL 4,5%)",   0.045),
    ]
    for r, lbl, pil_v in pil_scen:
        label_cell(ws, r, 1, lbl)
        input_cell(ws, r, 2, pil_v, fmt="0.0%")
        # Premio fed. = MAX(min, MIN(max, 40% * PIL_nom_medio))
        output_cell(ws, r, 3, f"=MAX($B$8,MIN($B$9,0.4*B{r}))", fmt="0.00%")
        # Cedole nette 12y: struttura step-up 4+4+4 (4a emissione MEF nov 2021)
        output_cell(ws, r, 4, f"=$B$12*(4*$B$5+4*$B$6+4*$B$7)*(1-$B$11)", fmt="#,##0")
        # Premio netto + cedole nette
        output_cell(ws, r, 5, f"=D{r}+$B$12*C{r}*(1-$B$11)", fmt="#,##0")
        # IRR nominale netto LINEARE = tot netto / capitale / durata (12y)
        output_cell(ws, r, 6, f"=E{r}/$B$12/$B$10", fmt="0.00%")

    section_header(ws, 20, "NOTA: BTP Futura NON e' piu' emesso dal 2021 (ricostruzione storica = contesto)", span=SPAN)
    note_cell(ws, 21, 1, "La tabella sopra RICOSTRUISCE il collocamento nov 2021 (chi sottoscrisse a 100 con premio PIL). "
                          "Struttura step-up 4+4+4 anni DEFINITIVA 4a emissione (1-4: 0,75%, 5-8: 1,35%, 9-12: 1,70%). "
                          "Per chi compra OGGI vale la sezione qui sotto: prezzo MOT 86,15 e NIENTE premio PIL.", span=SPAN)

    # ==== Comprato OGGI sul MOT: prezzo 86,15 → YTM lordo ~3,65% ====
    section_header(ws, 23, "Comprato OGGI sul MOT - prezzo 86,15 → YTM lordo ~3,65%", span=SPAN)
    label_cell(ws, 24, 1, "Prezzo MOT (uff. 09/06/2026)")
    output_cell(ws, 24, 2, "='1 - Parametri'!B66", fmt="0.00")
    label_cell(ws, 25, 1, "Anni residui (scad. 16/11/2033)")
    output_cell(ws, 25, 2, "='1 - Parametri'!D66", fmt="0.00")
    label_cell(ws, 26, 1, "Cedola media residua (step pesati sugli anni rimasti)")
    # Fasi ancorate alla scadenza: ultimi 4y = step3, 4y prima = step2, il resto = step1
    output_cell(ws, 26, 2, "=(B5*MAX(0,B25-8)+B6*MIN(4,MAX(0,B25-4))+B7*MIN(4,B25))/B25", fmt="0.00%")
    label_cell(ws, 27, 1, "YTM lordo (formula bond equivalent yield)")
    # BEY: (cedola_media + (100-P)/N) / ((100+P)/2) — stessa convenzione del foglio Parametri B50/B52
    output_cell(ws, 27, 2, "=(B26*100+(100-B24)/B25)/((100+B24)/2)", fmt="0.00%")
    label_cell(ws, 28, 1, "YTM netto (tassazione 12,5%)")
    output_cell(ws, 28, 2, "=B27*(1-B11)", fmt="0.00%")
    note_cell(ws, 29, 1, "Il forte sconto sotto par (86,15) compensa le cedole basse: quasi tutto il rendimento "
                          "arriva dal capital gain a scadenza (100-86,15 su 7,4 anni). "
                          "ATTENZIONE: il doppio premio PIL spetta SOLO ai sottoscrittori originali del nov 2021, "
                          "NON a chi compra sul MOT — il YTM qui sopra correttamente lo esclude.", span=SPAN)

    section_header(ws, 31, "IRR reale netto per scenario (da YTM MOT)", span=SPAN)
    table_header(ws, 32, 1, "Scenario inflazione")
    table_header(ws, 32, 2, "Infl. FOI media")
    table_header(ws, 32, 3, "IRR reale netto (YTM netto - FOI)")
    mot_scen = [
        (33, "1 - Disinflazione 1,0%",  "='1 - Parametri'!B37"),
        (34, "2 - Base BCE 2,0%",       "='1 - Parametri'!B38"),
        (35, "3 - Persistente 3,0%",    "='1 - Parametri'!B39"),
        (36, "4 - Stress geopol. 4,5%", "='1 - Parametri'!B40"),
    ]
    for r, lbl, ref in mot_scen:
        label_cell(ws, r, 1, lbl)
        output_cell(ws, r, 2, ref, fmt="0.0%")
        # Cedola nominale fissa: reale LINEARE = ytm_netto - FOI (convenzione foglio 7)
        output_cell(ws, r, 3, f"=$B$28-B{r}", fmt="0.00%")
    color_rule_fut_mot = ColorScaleRule(
        start_type='min', start_color='F87171',
        mid_type='percentile', mid_value=50, mid_color='FEF3C7',
        end_type='max', end_color='86EFAC',
    )
    ws.conditional_formatting.add('C33:C36', color_rule_fut_mot)

    # Bar chart: Tot netto 12y per scenario PIL (col E, rows 15-18)
    add_bar_chart(
        ws,
        title="BTP Futura - tot. netto 12 anni per scenario PIL",
        data_range=(5, 15, 5, 18),  # E15:E18 (header E15)
        cat_range=(1, 16, 1, 18),   # A16:A18
        anchor="H4",  # tabelle fino a col F: lascia G vuota di distacco
        y_title="EUR su 10k",
        y_fmt="#,##0",
    )
    # Heatmap su Tot netto 12y (più alto = meglio)
    color_rule_fut = ColorScaleRule(
        start_type='min', start_color='F87171',
        mid_type='percentile', mid_value=50, mid_color='FEF3C7',
        end_type='max', end_color='86EFAC',
    )
    ws.conditional_formatting.add('E16:E18', color_rule_fut)
    return ws


def build_btp_nominali(wb):
    """Foglio 6 - BTP nominali (2y/5y/10y curva)."""
    ws = wb.create_sheet("6 - BTP nominali")
    set_col_widths(ws, [42, 16, 16, 16, 16])
    SPAN = 5  # larghezza UNIFORME bande/sezioni = max colonna tabelle del foglio (col E)

    title_row(ws, 1, "BTP nominali - curva tassi 2026 (cedola fissa, NO inflazione)", span=SPAN)
    disclaimer_row(ws, 2, span=SPAN)

    section_header(ws, 4, "Input curva (da foglio Parametri)", span=SPAN)
    rows_in = [
        ("2 anni (feb 2028)",  "='1 - Parametri'!B30", 2),
        ("3 anni (mar 2029)",  "='1 - Parametri'!B31", 3),
        ("5 anni (stima)",     "='1 - Parametri'!B32", 5),
        ("7 anni (mar 2033)",  "='1 - Parametri'!B33", 7),
        ("10 anni (feb 2036)", "='1 - Parametri'!B34", 10),
    ]
    table_header(ws, 5, 1, "Scadenza")
    table_header(ws, 5, 2, "Rendimento lordo")
    table_header(ws, 5, 3, "Cedola netta annua")
    table_header(ws, 5, 4, "Tot. cedole nette a scad.")
    table_header(ws, 5, 5, "Cedola netta annua EUR")
    label_cell(ws, 6, 1, "Tassazione 12,5% applicata")
    label_cell(ws, 7, 1, "Capitale (default Parametri)")
    output_cell(ws, 6, 2, "='1 - Parametri'!B8", fmt="0.0%")
    output_cell(ws, 7, 2, "='1 - Parametri'!B43", fmt="#,##0")

    for i, (lbl, ref, n) in enumerate(rows_in):
        r = 9 + i
        label_cell(ws, r, 1, lbl)
        output_cell(ws, r, 2, ref, fmt="0.00%")
        output_cell(ws, r, 3, f"=B{r}*(1-$B$6)", fmt="0.000%")
        output_cell(ws, r, 4, f"=$B$7*C{r}*{n}", fmt="#,##0")
        # Cedola netta annua in EUR sul capitale
        output_cell(ws, r, 5, f"=$B$7*C{r}", fmt="#,##0")

    section_header(ws, 16, "NOTA: BTP nominali NON proteggono dall'inflazione", span=SPAN)
    note_cell(ws, 17, 1, "Rendimento certo in nominale. Se inflazione > rendimento, perdi potere d'acquisto. "
                          "Vantaggio: tassi attualmente alti, e plusvalenza se i tassi BCE scendono.", span=SPAN)

    # Aggiungo cella header dedicata per il chart in B8 (riga vuota gia' presente)
    table_header(ws, 8, 2, "Rendim. lordo (curva)")

    # Line chart: curva tassi 2y/3y/5y/7y/10y. Solo B8 (header) + B9-B13 (5 valori)
    add_line_chart(
        ws,
        title="Curva BTP nominali 2026 - rendimento lordo per scadenza",
        data_range=(2, 8, 2, 13),   # B8:B13 (header B8 + 5 valori soli rendimenti)
        cat_range=(1, 9, 1, 13),    # A9:A13 (scadenze)
        anchor="G4",
        y_title="Rendimento lordo",
        x_title="Scadenza",
    )
    return ws


def build_matrice(wb):
    """Foglio 7 - MATRICE COMPARATIVA 6 strumenti x 4 scenari inflazione (rendimenti REALI netti)."""
    ws = wb.create_sheet("7 - MATRICE COMPARATIVA")
    set_col_widths(ws, [38, 16, 16, 16, 16])
    SPAN = 5  # larghezza UNIFORME bande/sezioni = max colonna tabelle del foglio (col E)

    title_row(ws, 1, "MATRICE: rendimento REALE netto annuo - 6 strumenti x 4 scenari", span=SPAN)
    disclaimer_row(ws, 2, span=SPAN)

    section_header(ws, 4, "Tutti i valori sono IRR REALI annui netti (lordo - tasse - inflazione)", span=SPAN)
    table_header(ws, 5, 1, "Strumento")
    table_header(ws, 5, 2, "Infl. 1,0%")
    table_header(ws, 5, 3, "Infl. 2,0%")
    table_header(ws, 5, 4, "Infl. 3,0%")
    table_header(ws, 5, 5, "Infl. 4,5%")

    # Tasso fisso Si (es. 1,2%) e' GIA' REALE -> rendimento reale = tasso fisso + premio annualizzato (cedola tassata)
    # Per gli altri (nominali fissi): rendimento reale LINEARE = nom_netto - infl (niente Fisher)
    # Valore/Futura: YTM lordo a prezzo MOT (fogli 4/5, formula BEY — niente premi: non spettano a chi compra oggi)
    valore_ytm = "'4 - BTP Valore'!$B$32"   # YTM lordo MOT (~3,48% a 98,65)
    futura_ytm = "'5 - BTP Futura'!$B$27"   # YTM lordo MOT (~3,65% a 86,15)
    nominale_5y = "'1 - Parametri'!B32"
    # Classico: usa YTM REALE EFFETTIVO (media mar 2028 + giu 2030) - tiene conto prezzo MOT
    classico_real = "'1 - Parametri'!B56"  # YTM reale medio classico (~1,0% con prezzi MOT 09/06)
    si_real = "'1 - Parametri'!B5"
    tasse = "'1 - Parametri'!B8"
    premio_si_an = "('1 - Parametri'!B7)/5"  # 0,6%/5y = 0,12% annuo

    scenari = [("'1 - Parametri'!B37", "B37"),
               ("'1 - Parametri'!B38", "B38"),
               ("'1 - Parametri'!B39", "B39"),
               ("'1 - Parametri'!B40", "B40")]

    # Riga 6: BTP Italia Si - IRR REALE netto DIPENDE dallo scenario FOI:
    # cedola lorda = fisso + premio/durata + FOI; netta = *(1-tasse); reale = - FOI
    # => (fisso + premio/5 + FOI)*(1-t) - FOI  ==  (fisso + premio/5)*(1-t) - t*FOI
    # (drag fiscale: le tasse colpiscono anche la componente inflazione della cedola)
    label_cell(ws, 6, 1, "BTP Italia Si (1,2% + FOI)")
    for i, (sc, _) in enumerate(scenari):
        col = 2 + i
        # ROUND a 6 dp = tie deterministico con foglio 2 col F (S2 -> 0,91%)
        formula = f"=ROUND(({si_real}+{premio_si_an}+{sc})*(1-{tasse})-{sc},6)"
        output_cell(ws, 6, col, formula, fmt="0.00%")

    # Riga 7: BTP Italia classico - usa YTM REALE effettivo (chi compra OGGI sul MOT).
    # Anche il classico e' indicizzato FOI => stesso drag fiscale -t*FOI per scenario:
    # reale netto = ytm_medio_netto - tasse*FOI
    label_cell(ws, 7, 1, "BTP Italia classico (YTM reale MOT)")
    for i, (sc, _) in enumerate(scenari):
        col = 2 + i
        formula = f"={classico_real}*(1-{tasse})-{tasse}*{sc}"
        output_cell(ws, 7, col, formula, fmt="0.00%")

    # Riga 8: BTP Valore comprato OGGI sul MOT (YTM lordo BEY dal foglio 4, prezzo 98,65;
    # niente premio fedelta' per chi compra al secondario) - reale LINEARE: ytm*(1-tasse) - infl
    label_cell(ws, 8, 1, "BTP Valore (YTM lordo MOT ~3,48%)")
    for i, (sc, _) in enumerate(scenari):
        col = 2 + i
        formula = f"={valore_ytm}*(1-{tasse})-{sc}"
        output_cell(ws, 8, col, formula, fmt="0.00%")

    # Riga 9: BTP Futura comprato OGGI sul MOT (YTM lordo BEY dal foglio 5, prezzo 86,15;
    # niente premio PIL per chi compra al secondario) - reale LINEARE: ytm*(1-tasse) - infl
    label_cell(ws, 9, 1, "BTP Futura (YTM lordo MOT ~3,65%)")
    for i, (sc, _) in enumerate(scenari):
        col = 2 + i
        formula = f"={futura_ytm}*(1-{tasse})-{sc}"
        output_cell(ws, 9, col, formula, fmt="0.00%")

    # Riga 10: BTP nominale 5y - reale LINEARE: x*(1-tasse) - infl
    label_cell(ws, 10, 1, "BTP nominale 5y (2,90% lordo)")
    for i, (sc, _) in enumerate(scenari):
        col = 2 + i
        formula = f"={nominale_5y}*(1-{tasse})-{sc}"
        output_cell(ws, 10, col, formula, fmt="0.00%")

    # Riga 11: BTP nominale 10y - reale LINEARE: x*(1-tasse) - infl
    nominale_10y = "'1 - Parametri'!B34"
    label_cell(ws, 11, 1, "BTP nominale 10y (3,45% lordo)")
    for i, (sc, _) in enumerate(scenari):
        col = 2 + i
        formula = f"={nominale_10y}*(1-{tasse})-{sc}"
        output_cell(ws, 11, col, formula, fmt="0.00%")

    section_header(ws, 13, "Heatmap: VERDE = miglior rendimento REALE per scenario / ROSSO = peggior", span=SPAN)
    note_cell(ws, 14, 1, "BTP Italia Si vince in scenari di inflazione media-alta perche' il tasso fisso e' REALE. "
                          "BTP Italia classico in inflazione alta ha vantaggio capitale rivalutato. "
                          "Valore e Futura: YTM da prezzo MOT (comprati OGGI, senza premi fedelta'/PIL), "
                          "cedola nominale fissa → vincono solo con inflazione bassa. "
                          "Nominali: vincono solo in disinflazione.", span=SPAN)

    # Conditional formatting: colorscale rosso-giallo-verde sulle 4 colonne dei valori (B6:E11)
    # Applica a OGNI colonna separatamente (= confronto strumenti per quello scenario)
    color_rule = ColorScaleRule(
        start_type='min', start_color='F87171',     # rosso
        mid_type='percentile', mid_value=50, mid_color='FEF3C7',  # giallo chiaro
        end_type='max', end_color='86EFAC',         # verde
    )
    for col_letter in ['B', 'C', 'D', 'E']:
        ws.conditional_formatting.add(f'{col_letter}6:{col_letter}11', color_rule)

    # Line chart KILLER: matrice 6 strumenti x 4 scenari
    ch = LineChart()
    ch.style = 12
    ch.title = "Matrice: rendimento REALE netto annuo per strumento × inflazione"
    ch.y_axis.title = "Rendim. reale netto"
    ch.x_axis.title = "Scenario inflazione FOI"
    data = Reference(ws, min_col=1, min_row=6, max_col=5, max_row=11)
    cats = Reference(ws, min_col=2, min_row=5, max_col=5, max_row=5)
    ch.add_data(data, titles_from_data=True, from_rows=True)
    ch.set_categories(cats)
    ch.width = 26
    ch.height = 15
    # Colori distinti per le 6 linee (BTP Sì, classico, Valore, Futura, nom 5y, nom 10y)
    palette = ["10B981", "1E40AF", "F59E0B", "0EA5E9", "EF4444", "8B5CF6"]
    for i, ser in enumerate(ch.series):
        if i < len(palette):
            gp_line = GraphicalProperties()
            gp_line.line = LineProperties(solidFill=palette[i], w=28000)
            ser.graphicalProperties = gp_line
    # Chart multi-serie: plot_w piu' stretto per spazio legenda dx + plot_y top per title
    _style_axes(ch, y_fmt="0.00%", plot_x=0.16, plot_y=0.18, plot_w=0.64, plot_h=0.62)
    ws.add_chart(ch, "G4")
    return ws


def build_break_even(wb):
    """Foglio 8 - Break-even multi-strumento."""
    ws = wb.create_sheet("8 - Break-even")
    set_col_widths(ws, [44, 18, 16, 16, 22])
    SPAN = 5  # larghezza UNIFORME bande/sezioni = max colonna tabelle del foglio (col E)

    title_row(ws, 1, "Break-even: a quale inflazione media il BTP Italia Si batte ogni avversario", span=SPAN)
    disclaimer_row(ws, 2, span=SPAN)

    section_header(ws, 4, "Avversari a cedola FISSA nominale (break-even basato su inflazione)", span=SPAN)
    table_header(ws, 5, 1, "Avversario")
    table_header(ws, 5, 2, "Rendimento lordo")
    table_header(ws, 5, 3, "Rend. netto (1-12,5%)")
    table_header(ws, 5, 4, "Soglia infl. break-even (lorda)")
    table_header(ws, 5, 5, "Verdetto")

    # Break-even Si vs nominale: stessa tassazione 12,5% su entrambi i lati
    # => le tasse si ELIDONO: FOI_BE = nom_LORDO - (fisso_si + premio_si/durata)
    # (NON moltiplicare per (1-t): la soglia e' una differenza di rendimenti LORDI)
    si_lordo_an = "('1 - Parametri'!B5+('1 - Parametri'!B7)/5)"
    tasse = "'1 - Parametri'!B8"

    # Valore/Futura: YTM lordo a prezzo MOT (formula BEY, fogli 4/5) — comprati OGGI al
    # secondario, quindi NIENTE premio fedelta'/PIL (spettano solo ai sottoscrittori originali).
    avversari = [
        (6, "BTP nominale 2y",   "='1 - Parametri'!B30"),
        (7, "BTP nominale 3y",   "='1 - Parametri'!B31"),
        (8, "BTP nominale 5y",   "='1 - Parametri'!B32"),
        (9, "BTP nominale 7y",   "='1 - Parametri'!B33"),
        (10,"BTP nominale 10y",  "='1 - Parametri'!B34"),
        (11,"BTP Valore (YTM lordo MOT 98,65)",  "='4 - BTP Valore'!B32"),
        (12,"BTP Futura (YTM lordo MOT 86,15)",  "='5 - BTP Futura'!B27"),
    ]
    for r, lbl, ref in avversari:
        label_cell(ws, r, 1, lbl)
        output_cell(ws, r, 2, ref, fmt="0.00%")
        # Col C: rendimento netto SOLO informativo (non entra nel break-even)
        output_cell(ws, r, 3, f"=B{r}*(1-{tasse})", fmt="0.00%")
        # Col D: soglia LORDA (vedi commento sopra) - usata da verdetto IF e bar chart
        output_cell(ws, r, 4, f"=B{r}-{si_lordo_an}", fmt="0.00%")
        # Verdetto: ROUND*100&"%" per locale-safe (2 decimali per soglie precise)
        output_cell(ws, r, 5, f'=IF(D{r}<=0,"Si vince SEMPRE",IF(D{r}>=0.045,"Si perde quasi sempre","Si vince se FOI > "&ROUND(D{r}*100,2)&"%"))', fmt="@")

    # BTP Italia classico: confronto diverso (rivaluta capitale, no break-even semplice)
    # Confronto a parita' di inflazione su base REALE NETTA: entrambi indicizzati FOI,
    # il drag fiscale -t*FOI si elide => basta confrontare le componenti reali nette.
    section_header(ws, 14, "BTP Italia classico: confronto diverso (capitale rivalutato, no break-even classico)", span=SPAN)
    table_header(ws, 15, 1, "Avversario")
    table_header(ws, 15, 2, "YTM reale NETTO (MOT)")
    table_header(ws, 15, 3, "Sì reale netto")
    table_header(ws, 15, 4, "Vantaggio Sì (pp)")
    table_header(ws, 15, 5, "Verdetto")

    si_reale_netto = "=('1 - Parametri'!B5+'1 - Parametri'!B7/5)*(1-'1 - Parametri'!B8)"

    label_cell(ws, 16, 1, "BTP Italia classico mar 2028 (YTM REALE da prezzo MOT)")
    output_cell(ws, 16, 2, "='1 - Parametri'!B50*(1-'1 - Parametri'!B8)", fmt="0.00%")
    output_cell(ws, 16, 3, si_reale_netto, fmt="0.00%")
    output_cell(ws, 16, 4, "=C16-B16", fmt="0.00%")
    output_cell(ws, 16, 5, '=IF(D16>=0,"Sì vince (vantaggio "&ROUND(D16*100,2)&"pp)","Classico vince (gap "&ROUND(-D16*100,2)&"pp)")', fmt="@")

    label_cell(ws, 17, 1, "BTP Italia classico giu 2030 (YTM REALE da prezzo MOT)")
    output_cell(ws, 17, 2, "='1 - Parametri'!B52*(1-'1 - Parametri'!B8)", fmt="0.00%")
    output_cell(ws, 17, 3, si_reale_netto, fmt="0.00%")
    output_cell(ws, 17, 4, "=C17-B17", fmt="0.00%")
    output_cell(ws, 17, 5, '=IF(D17>=0,"Sì vince (vantaggio "&ROUND(D17*100,2)&"pp)","Classico vince (gap "&ROUND(-D17*100,2)&"pp)")', fmt="@")

    label_cell(ws, 18, 1, "BTP Italia classico MEDIA 2 ISIN (YTM reale NETTO medio)")
    output_cell(ws, 18, 2, "='1 - Parametri'!B56*(1-'1 - Parametri'!B8)", fmt="0.00%")
    output_cell(ws, 18, 3, si_reale_netto, fmt="0.00%")
    output_cell(ws, 18, 4, "=C18-B18", fmt="0.00%")
    output_cell(ws, 18, 5, '=IF(D18>=0,"Sì vince (vantaggio "&ROUND(D18*100,2)&"pp)","Classico vince (gap "&ROUND(-D18*100,2)&"pp)")', fmt="@")

    note_cell(ws, 19, 1, "ATTENZIONE: usato YTM REALE EFFETTIVO calcolato dal prezzo MOT (non cedola del prospetto). "
                          "Se sul secondario il classico quota SOPRA par, il YTM reale si abbassa, specie per scadenze "
                          "brevi. Esempio coi prezzi uff. 09/06/2026: mar 2028 a 101,97 → YTM reale ~0,9%; giu 2030 a "
                          "101,92 → ~1,1%. Vai al foglio Parametri righe 46-56 per modificare i prezzi MOT.", span=SPAN)

    note_cell(ws, 21, 1, "BTP Futura comprato OGGI sul MOT a 86,15 (forte sconto sotto par): YTM lordo ~3,65% "
                          "nonostante le cedole step-up basse (0,75/1,35/1,70). Soglia break-even FOI ~2,3%: "
                          "sotto, il Futura vince; sopra, vince il Sì. Il doppio premio PIL NON spetta a chi compra "
                          "sul MOT (solo sottoscrittori originali nov 2021).", span=SPAN)

    # Bar chart orizzontale: soglia inflazione break-even per ogni avversario fisso-nominale
    add_bar_chart(
        ws,
        title="Break-even: soglia inflazione FOI per ogni avversario (cedola fissa nominale)",
        data_range=(4, 5, 4, 12),   # D5:D12 (header D5 + 7 avversari ora con Futura)
        cat_range=(1, 6, 1, 12),    # A6:A12
        anchor="G4",
        y_title="Soglia inflazione",
        bar_dir="bar",  # orizzontale
    )
    # Heatmap soglia inflazione INVERSA: bassa = facile per Sì (verde), alta = difficile (rosso)
    color_rule_be = ColorScaleRule(
        start_type='min', start_color='86EFAC',   # verde: soglia bassa = Sì vince facile
        mid_type='percentile', mid_value=50, mid_color='FEF3C7',
        end_type='max', end_color='F87171',       # rosso: soglia alta = Sì fatica
    )
    ws.conditional_formatting.add('D6:D12', color_rule_be)
    # Heatmap vantaggio Sì (alto = meglio per Sì = verde)
    color_rule_gap = ColorScaleRule(
        start_type='min', start_color='F87171',
        mid_type='percentile', mid_value=50, mid_color='FEF3C7',
        end_type='max', end_color='86EFAC',
    )
    ws.conditional_formatting.add('D16:D18', color_rule_gap)
    return ws


def build_profili(wb):
    """Foglio 9 - 4 profili FIRE + allocazione consigliata."""
    ws = wb.create_sheet("9 - 4 profili FIRE")
    set_col_widths(ws, [34, 14, 14, 14, 14, 14])
    SPAN = 6  # larghezza UNIFORME bande/sezioni = max colonna tabelle del foglio (col F)

    title_row(ws, 1, "4 profili FIRE - allocazione consigliata tra i 5 strumenti", span=SPAN)
    disclaimer_row(ws, 2, span=SPAN)

    section_header(ws, 4, "Allocazione (% del cuscinetto bond del portafoglio)", span=SPAN)
    table_header(ws, 5, 1, "Profilo")
    table_header(ws, 5, 2, "BTP Italia Si")
    table_header(ws, 5, 3, "BTP Italia clas.")
    table_header(ws, 5, 4, "BTP Valore")
    table_header(ws, 5, 5, "BTP nominale 5y")
    table_header(ws, 5, 6, "BTP nominale 10y")

    profili = [
        (6, "Giovane 25-34",     0.20, 0.10, 0.20, 0.30, 0.20),
        (7, "Mezza età 35-49",   0.25, 0.15, 0.20, 0.25, 0.15),
        (8, "Over 50 pre-FIRE",  0.30, 0.30, 0.20, 0.15, 0.05),
        (9, "FIRE in decumulo",  0.25, 0.35, 0.15, 0.15, 0.10),
    ]
    for r, lbl, a, b, c, d, e in profili:
        label_cell(ws, r, 1, lbl)
        output_cell(ws, r, 2, a, fmt="0%")
        output_cell(ws, r, 3, b, fmt="0%")
        output_cell(ws, r, 4, c, fmt="0%")
        output_cell(ws, r, 5, d, fmt="0%")
        output_cell(ws, r, 6, e, fmt="0%")

    section_header(ws, 11, "Razionale per profilo", span=SPAN)
    razionali = [
        (12, "Giovane: peso nominali alto (tassi sopra inflazione attesa), Si come 'difesa' inflazione."),
        (13, "Mezza eta': bilanciato, Si cresce per protezione potere d'acquisto."),
        (14, "Over 50: peso classico alto per crescita capitale; Si per cedola decrescente in real terms."),
        (15, "Decumulo: priorita' capitale rivalutato (classico) + reddito stabile (Si + Valore)."),
    ]
    for r, txt in razionali:
        note_cell(ws, r, 1, txt, span=SPAN)

    # Heatmap allocazione % per profilo (più alto = forte presenza in portafoglio)
    color_rule_prof = ColorScaleRule(
        start_type='min', start_color='F87171',
        mid_type='percentile', mid_value=50, mid_color='FEF3C7',
        end_type='max', end_color='86EFAC',
    )
    ws.conditional_formatting.add('B6:F9', color_rule_prof)

    # Stacked bar chart: allocazione % per profilo (5 strumenti impilati)
    ch = BarChart()
    ch.type = "col"
    ch.style = 11
    ch.grouping = "percentStacked"
    ch.overlap = 100
    ch.title = "Allocazione % per profilo FIRE (5 strumenti)"
    ch.y_axis.title = "% portafoglio bond"
    data = Reference(ws, min_col=2, min_row=5, max_col=6, max_row=9)
    cats = Reference(ws, min_col=1, min_row=6, max_col=1, max_row=9)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ch.width = 22
    ch.height = 13
    # Colori distinti per le 5 serie (BTP Sì, classico, Valore, nom 5y, nom 10y)
    palette = ["10B981", "1E40AF", "F59E0B", "EF4444", "8B5CF6"]  # emerald, blu, amber, rosso, viola
    for i, ser in enumerate(ch.series):
        if i < len(palette):
            ser.graphicalProperties = GraphicalProperties(solidFill=palette[i])
    # Stacked multi-serie: plot_w stretto per legenda dx + plot_y top per title
    _style_axes(ch, y_fmt="0%", plot_x=0.16, plot_y=0.18, plot_w=0.64, plot_h=0.62)
    ws.add_chart(ch, "H4")
    return ws


def build_calcolatore(wb):
    """Foglio 10 - calcolatore personale."""
    ws = wb.create_sheet("10 - Calcolatore personale")
    set_col_widths(ws, [42, 16, 30])
    SPAN = 3  # larghezza UNIFORME bande/sezioni = max colonna tabelle del foglio (col C)

    title_row(ws, 1, "Calcolatore: quale BTP scegliere in base ai tuoi parametri", span=SPAN)
    disclaimer_row(ws, 2, span=SPAN)

    section_header(ws, 4, "Input personali", span=SPAN)
    inputs = [
        ("Tuo capitale (EUR)",                          20000, "#,##0"),
        ("Tuo orizzonte temporale (anni)",              7,     "0"),
        ("Tua aspettativa inflazione FOI media (%)",    0.025, "0.0%"),
        ("Tua aliquota IRPEF marginale (info)",         0.33,  "0%"),
    ]
    for i, (lbl, val, fmt) in enumerate(inputs, start=5):
        label_cell(ws, i, 1, lbl)
        input_cell(ws, i, 2, val, fmt=fmt)

    section_header(ws, 10, "Rendimento REALE netto atteso per strumento", span=SPAN)
    table_header(ws, 11, 1, "Strumento")
    table_header(ws, 11, 2, "Rend. reale annuo")
    table_header(ws, 11, 3, "Idoneo al tuo orizzonte")
    tasse = "'1 - Parametri'!B8"
    infl_user = "B7"

    # Convenzioni COERENTI col foglio 7 (tutto LINEARE):
    # - Indicizzati FOI (Sì, classico): reale netto = reale_netto_base - tasse*infl (drag fiscale su comp. inflazione)
    # - Nominali (Valore, 5y, 10y): reale netto = nominale_netto - infl (NO drag: cedola non indicizzata)
    # - Valore: YTM lordo a prezzo MOT (foglio 4 B32) — chi compra OGGI non ha premio fedelta'
    strumenti = [
        (12, "BTP Italia Si",    "=('1 - Parametri'!B5+'1 - Parametri'!B7/5)*(1-{tasse})-{tasse}*{infl_user}", 5),
        (13, "BTP Italia clas. (YTM MOT)", "='1 - Parametri'!B56*(1-{tasse})-{tasse}*{infl_user}", 5),
        (14, "BTP Valore (YTM MOT)", "='4 - BTP Valore'!B32*(1-{tasse})-{infl_user}", 6),
        (15, "BTP nominale 5y",  "='1 - Parametri'!B32*(1-{tasse})-{infl_user}", 5),
        (16, "BTP nominale 10y", "='1 - Parametri'!B34*(1-{tasse})-{infl_user}", 10),
    ]
    for r, lbl, formula, dur in strumenti:
        label_cell(ws, r, 1, lbl)
        f = formula.format(tasse=tasse, infl_user=infl_user)
        output_cell(ws, r, 2, f, fmt="0.00%")
        output_cell(ws, r, 3, f'=IF(B6>={dur},"SI","NO - troppo lungo")', fmt="@")

    section_header(ws, 18, "Raccomandazione automatica (strumento con rendimento reale max)", span=SPAN)
    label_cell(ws, 19, 1, "Miglior strumento")
    output_cell(ws, 19, 2, "=INDEX(A12:A16,MATCH(MAX(B12:B16),B12:B16,0))", fmt="@")
    label_cell(ws, 20, 1, "Rendimento reale atteso annuo")
    output_cell(ws, 20, 2, "=MAX(B12:B16)", fmt="0.00%")
    label_cell(ws, 21, 1, "Valore finale (con cap. + reale)")
    output_cell(ws, 21, 2, "=B5*(1+B20)^B6", fmt="#,##0")

    note_cell(ws, 23, 1, "Raccomandazione algoritmica basata sul rendimento reale max. "
                          "NON e' consulenza finanziaria. Considera anche liquidita', diversificazione, "
                          "necessita' di reddito cedolare periodico.", span=SPAN)

    # Bar chart: rendimento reale annuo per strumento (col B, rows 11-16)
    add_bar_chart(
        ws,
        title="Rendimento REALE netto annuo per strumento (tuoi input)",
        data_range=(2, 11, 2, 16),  # B11:B16 (header B11 + 5 strumenti)
        cat_range=(1, 12, 1, 16),   # A12:A16
        anchor="E4",
        y_title="Rendimento reale annuo",
    )
    # Heatmap rendimento reale per scelta consigliata
    color_rule_calc = ColorScaleRule(
        start_type='min', start_color='F87171',
        mid_type='percentile', mid_value=50, mid_color='FEF3C7',
        end_type='max', end_color='86EFAC',
    )
    ws.conditional_formatting.add('B12:B16', color_rule_calc)
    return ws


def main(out_path: str | None = None):
    wb = Workbook()

    # Cover info
    title = "Simulatore BTP Italia Si vs altri BTP retail (giugno 2026)"
    sheets_info = [
        ("1 - Parametri",        "Input centralizzati (modifica QUI dopo annuncio MEF 12 giu)"),
        ("2 - BTP Italia Si",     "Rendimento Si in 4 scenari inflazione FOI"),
        ("3 - BTP Italia classico","Capitale rivalutato + cedola reale 1,6% (giu 2030)"),
        ("4 - BTP Valore",        "Step-up 2,60/3,20/3,80 + YTM a prezzo MOT 98,65"),
        ("5 - BTP Futura",        "Storico 4a emissione + YTM a prezzo MOT 86,15"),
        ("6 - BTP nominali",      "Curva 2y/5y/10y rendimenti fissi"),
        ("7 - MATRICE COMPARATIVA","6 strumenti x 4 scenari (rendimenti REALI netti)"),
        ("8 - Break-even",        "Soglia inflazione per ogni avversario"),
        ("9 - 4 profili FIRE",    "Allocazione consigliata"),
        ("10 - Calcolatore",      "Tuo profilo personale - raccomandazione automatica"),
    ]
    sources = [
        "MEF Dipartimento del Tesoro - Annuncio BTP Italia Si dal 15 al 19 giugno 2026",
        "MEF - BTP Italia marzo 2028 (IT0005532723, scad. 14/03/2028) tasso reale 2,0%; giugno 2030 (IT0005497000) tasso reale 1,6%",
        "MEF - BTP Valore marzo 2026 (IT0005696338) step-up DEFINITIVI 2,60/3,20/3,80% + premio 0,8%, scadenza 10 mar 2032",
        "MEF - BTP Futura 4a emissione nov 2021 (IT0005466351), 12 anni step-up DEFINITIVI 0,75/1,35/1,70% + doppio premio PIL (solo sottoscrittori originali), scadenza 16 nov 2033",
        "Borsa Italiana MOT - prezzi ufficiali 09/06/2026: mar 2028 101,97; giu 2030 101,92; Valore 98,65; Futura 86,15",
        "MEF Tesoro - Calendario aste BTP nominali maggio 2026 (asta 13 e 28 maggio)",
        "Banca d'Italia - Risultati aste BTP nominali 2026 (rendimenti curva 2y-10y)",
        "ISTAT - Indice FOI per famiglie operai e impiegati (escluso tabacchi) - meccanismo BTP Italia",
        "Eurostat - HICP eurozona per BTPi (non incluso in matrice; alternativa istituzionale)",
        "Borsa Italiana MOT - mercato secondario BTP retail (liquidita' supportata da Dealer MEF)",
    ]
    build_cover(wb, title, sheets_info, sources)

    # Fogli operativi
    build_params(wb)
    build_btp_si(wb)
    build_btp_classico(wb)
    build_btp_valore(wb)
    build_btp_futura(wb)
    build_btp_nominali(wb)
    build_matrice(wb)
    build_break_even(wb)
    build_profili(wb)
    build_calcolatore(wb)

    # Output: di default nel repo STAGING private (pattern release_excel.py:
    # staging -> public solo al go-live via tools/release_excel.py --slug btp_si_compare).
    # MAI scrivere direttamente nel repo public da questo builder.
    if out_path is not None:
        out = Path(out_path)
    else:
        staging_root = Path("e:/sviluppo/marco-fire-simulatori-staging")
        if not staging_root.exists():
            sys.exit("FAIL: repo staging non trovato (e:/sviluppo/marco-fire-simulatori-staging). "
                     "Usa --out <path> per un output esplicito.")
        out = staging_root / "simulatori" / "btp_si_compare_2026.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"[OK] {out}  ({out.stat().st_size / 1024:.1f} KB)")


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--out", default=None,
                   help="override path output .xlsx (default: repo staging simulatori/)")
    args = p.parse_args()
    main(args.out)
